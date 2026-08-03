import itertools
import logging
import os
import re
from abc import ABCMeta
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from pgvector.django import HnswIndex, VectorField

from django.conf import settings
from django.core.files.storage import default_storage
from django.db import models, transaction
from django.db.models import F, Q, Sum, Value
from django.db.models.functions import Cast, Lower
from django.utils import timezone
from django.utils.text import slugify
from django.utils.translation import gettext_lazy as _

from temba import mailroom
from temba.contacts.models import Contact
from temba.orgs.models import DependencyMixin, Export, ExportType, Org, OrgMembership
from temba.users.models import User
from temba.utils.dates import date_range
from temba.utils.db.functions import SplitPart
from temba.utils.export import MultiSheetExporter
from temba.utils.models import TembaModel, delete_in_batches
from temba.utils.s3 import public_file_storage
from temba.utils.uuid import is_uuid, uuid4

logger = logging.getLogger(__name__)


class Shortcut(TembaModel):
    """
    A canned response available from the ticketing interface.
    """

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="shortcuts")
    text = models.TextField(max_length=10_000)

    @classmethod
    def create(cls, org, user, name: str, text: str):
        assert cls.is_valid_name(name), f"'{name}' is not a valid shortcut name"
        assert not org.shortcuts.filter(name__iexact=name).exists(), f"shortcut with name '{name}' already exists"

        return org.shortcuts.create(name=name, text=text, created_by=user, modified_by=user)

    def release(self, user):
        self.is_active = False
        self.name = self._deleted_name()
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_by", "modified_on"))

    class Meta:
        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_shortcut_names")]
        indexes = [
            # mailroom's staleness + delta sweep for the shortcuts knowledge source
            models.Index(name="shortcut_by_modified", fields=("org", "modified_on")),
        ]


class Knowledge(TembaModel):
    """
    A source of knowledge that AI agents can search semantically.

    Indexing - crawling, extracting, chunking and embedding - is performed entirely by mailroom, which sweeps for rows
    needing work. This app owns the schema, the CRUD UI and document uploads only; it never calls an embeddings service.
    """

    # authored types - items live in their own Django-owned table, mailroom only reads
    TYPE_SHORTCUTS = "shortcuts"  # the org's shortcuts, read by mailroom straight from tickets_shortcut
    TYPE_HELPDESK = "helpdesk"  # the org's own help articles, read from tickets_article
    # ingested types - items share tickets_knowledgeitem, mailroom owns the lifecycle
    TYPE_WEBSITE = "website"  # a crawled website
    TYPE_DOCUMENTS = "documents"  # uploaded files
    TYPE_CHOICES = (
        (TYPE_SHORTCUTS, _("Shortcuts")),
        (TYPE_HELPDESK, _("Helpdesk")),
        (TYPE_WEBSITE, _("Website")),
        (TYPE_DOCUMENTS, _("Documents")),
    )

    SYSTEM_TYPES = (TYPE_SHORTCUTS, TYPE_HELPDESK)  # one of each per org, created in Org.initialize()
    SYSTEM_NAMES = {TYPE_SHORTCUTS: "Shortcuts", TYPE_HELPDESK: "Helpdesk"}

    STATUS_PENDING = "P"  # needs (re)indexing, mailroom's sweep will pick it up
    STATUS_INDEXING = "I"  # mailroom is working on it
    STATUS_READY = "R"  # indexed and searchable
    STATUS_FAILED = "F"  # last indexing attempt failed, see error
    STATUS_CHOICES = (
        (STATUS_PENDING, _("Pending")),
        (STATUS_INDEXING, _("Indexing")),
        (STATUS_READY, _("Ready")),
        (STATUS_FAILED, _("Failed")),
    )

    REFRESH_NEVER = "never"
    REFRESH_DAILY = "daily"
    REFRESH_WEEKLY = "weekly"
    REFRESH_MONTHLY = "monthly"
    REFRESH_CHOICES = (
        (REFRESH_NEVER, _("Never")),
        (REFRESH_DAILY, _("Daily")),
        (REFRESH_WEEKLY, _("Weekly")),
        (REFRESH_MONTHLY, _("Monthly")),
    )

    # config keys for TYPE_WEBSITE
    CONFIG_URL = "url"
    CONFIG_MAX_DEPTH = "max_depth"
    CONFIG_MAX_PAGES = "max_pages"
    CONFIG_REFRESH = "refresh"

    DEFAULT_MAX_DEPTH = 3
    DEFAULT_MAX_PAGES = 500
    MAX_MAX_PAGES = 5_000
    MAX_URL_LEN = 2048

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="knowledge")
    knowledge_type = models.CharField(max_length=16, choices=TYPE_CHOICES)

    # type specific settings, e.g. for website: url, max_depth, max_pages, refresh. Empty for other types.
    config = models.JSONField(default=dict)

    # everything below is written by mailroom as it indexes - this app only reads it
    status = models.CharField(max_length=1, choices=STATUS_CHOICES, default=STATUS_PENDING)
    error = models.CharField(max_length=255, null=True)
    last_indexed_on = models.DateTimeField(null=True)
    num_items = models.IntegerField(default=0)
    num_chunks = models.IntegerField(default=0)

    org_limit_key = Org.LIMIT_KNOWLEDGE

    @classmethod
    def create_system(cls, org):
        """
        Creates the org's two system sources - its shortcut list and its helpdesk.
        """
        assert not org.knowledge.filter(knowledge_type__in=cls.SYSTEM_TYPES).exists(), (
            "org already has system knowledge"
        )

        return [
            org.knowledge.create(
                name=cls.SYSTEM_NAMES[t],
                knowledge_type=t,
                is_system=True,
                created_by=org.created_by,
                modified_by=org.modified_by,
            )
            for t in cls.SYSTEM_TYPES
        ]

    @classmethod
    def create_website(cls, org, user, name: str, url: str, *, max_depth=None, max_pages=None, refresh=None):
        assert cls.is_valid_name(name), f"'{name}' is not a valid knowledge name"
        assert not org.knowledge.filter(name__iexact=name, is_active=True).exists()

        return org.knowledge.create(
            name=name,
            knowledge_type=cls.TYPE_WEBSITE,
            config={
                cls.CONFIG_URL: url,
                cls.CONFIG_MAX_DEPTH: max_depth or cls.DEFAULT_MAX_DEPTH,
                cls.CONFIG_MAX_PAGES: max_pages or cls.DEFAULT_MAX_PAGES,
                cls.CONFIG_REFRESH: refresh or cls.REFRESH_WEEKLY,
            },
            created_by=user,
            modified_by=user,
        )

    @classmethod
    def create_documents(cls, org, user, name: str):
        assert cls.is_valid_name(name), f"'{name}' is not a valid knowledge name"
        assert not org.knowledge.filter(name__iexact=name, is_active=True).exists()

        # nothing to index until files are uploaded
        return org.knowledge.create(
            name=name,
            knowledge_type=cls.TYPE_DOCUMENTS,
            status=cls.STATUS_READY,
            created_by=user,
            modified_by=user,
        )

    @property
    def url(self) -> str:
        return self.config.get(self.CONFIG_URL)

    def mark_pending(self):
        """
        Flags this source as needing (re)indexing so mailroom's sweep picks it up. Called whenever this app changes
        something mailroom's index is derived from - website config, uploaded files.
        """
        self.status = self.STATUS_PENDING
        self.error = None
        self.save(update_fields=("status", "error"))

    def release(self, user):
        assert not (self.is_system and self.org.is_active), "can't release system knowledge"

        # deactivate first so nothing reads or indexes it while we purge
        self.is_active = False
        self.name = self._deleted_name()
        self.num_items = 0
        self.num_chunks = 0
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "num_items", "num_chunks", "modified_by", "modified_on"))

        self._purge()

    def delete(self):
        self._purge()

        super().delete()

    def _purge(self):
        """
        Removes this source's chunks, items, articles and article images. Rows go first; storage objects are only
        removed once their rows are gone.
        """
        # collect storage keys before the rows that name them disappear - two different buckets
        item_paths = list(self.items.exclude(path=None).values_list("path", flat=True))
        image_paths = list(ArticleImage.objects.filter(article__knowledge=self).values_list("path", flat=True))

        delete_in_batches(self.chunks.all())
        delete_in_batches(self.items.all())
        delete_in_batches(ArticleImage.objects.filter(article__knowledge=self))

        # parent is PROTECT so flatten the article tree before deleting it
        self.articles.exclude(parent=None).update(parent=None)
        delete_in_batches(self.articles.all())

        for path in item_paths:
            default_storage.delete(path)
        for path in image_paths:
            public_file_storage.delete(path)

    class Meta:
        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_knowledge_names")]
        indexes = [
            # mailroom's indexing sweep's worklist
            models.Index(name="knowledge_pending", fields=("id",), condition=Q(is_active=True, status="P")),
        ]


class Article(models.Model):
    """
    An article in an org's helpdesk. Written by this app, read by mailroom, which indexes only published, active
    articles.

    Deliberately not a TembaModel: TembaModel.name is capped at 64 chars and NameValidator rejects " and \\, which real
    help titles routinely contain. Soft-deleted like Shortcut so mailroom's delta sweep sees the tombstone - a hard
    delete would leave its chunks stranded until a full reindex.
    """

    STATUS_DRAFT = "D"  # never indexed, never public
    STATUS_PUBLISHED = "P"
    STATUS_CHOICES = ((STATUS_DRAFT, _("Draft")), (STATUS_PUBLISHED, _("Published")))

    MAX_TITLE_LEN = 255
    MAX_SLUG_LEN = 255
    MAX_DEPTH = 3  # root + two levels of nesting; enforced by the (phase 4) reorder view

    uuid = models.UUIDField(unique=True, default=uuid4)
    knowledge = models.ForeignKey(Knowledge, on_delete=models.PROTECT, related_name="articles")

    # the tree - a plain self-FK, not mptt (that dep exists only for locations and buys nothing at help-centre depth).
    # Depth is capped at MAX_DEPTH and cycles are rejected server-side.
    parent = models.ForeignKey("self", on_delete=models.PROTECT, null=True, related_name="children")
    sort_order = models.IntegerField(default=0)

    title = models.CharField(max_length=MAX_TITLE_LEN)
    slug = models.SlugField(max_length=MAX_SLUG_LEN)
    body = models.TextField(default="")  # markdown source

    status = models.CharField(max_length=1, choices=STATUS_CHOICES, default=STATUS_DRAFT)
    published_on = models.DateTimeField(null=True)

    is_active = models.BooleanField(default=True)
    created_by = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.PROTECT, related_name="+")
    created_on = models.DateTimeField(default=timezone.now)
    modified_by = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.PROTECT, related_name="+")
    # auto_now is load-bearing: mailroom's staleness sweep is MAX(modified_on) > knowledge.last_indexed_on, so an
    # unpublish or a soft-delete has to bump it for the removal to be noticed
    modified_on = models.DateTimeField(auto_now=True)

    @classmethod
    def get_unique_slug(cls, knowledge, title: str, ignore=None) -> str:
        base = slugify(title)[: cls.MAX_SLUG_LEN] or "article"
        qs = cls.objects.filter(knowledge=knowledge, is_active=True)
        if ignore:
            qs = qs.exclude(id=ignore.id)

        slug, count = base, 1
        while qs.filter(slug=slug).exists():
            count += 1
            suffix = f"-{count}"
            slug = f"{base[: cls.MAX_SLUG_LEN - len(suffix)]}{suffix}"

        return slug

    @property
    def org(self):
        return self.knowledge.org

    def release(self, user):
        """
        Soft delete. Children are reparented to our parent so the tree stays connected.
        """
        self.children.update(parent=self.parent)

        self.is_active = False
        self.status = self.STATUS_DRAFT
        self.modified_by = user
        self.save(update_fields=("is_active", "status", "modified_by", "modified_on"))

    class Meta:
        constraints = [
            models.UniqueConstraint("knowledge", "slug", condition=Q(is_active=True), name="unique_article_slugs")
        ]
        indexes = [
            # the tree, in display order
            models.Index(name="article_by_tree", fields=("knowledge", "parent", "sort_order")),
            # mailroom's staleness + delta sweep
            models.Index(name="article_by_modified", fields=("knowledge", "modified_on")),
        ]


def get_article_image_path(article, image_uuid, filename: str) -> str:
    return (
        f"orgs/{article.knowledge.org_id}/knowledge/{article.knowledge.uuid}/"
        f"articles/{article.uuid}/{image_uuid}{Path(filename).suffix.lower()}"
    )


class ArticleImage(models.Model):
    """
    A screenshot uploaded to an article and referenced from its markdown by URL. Stored in public storage because the
    eventual standalone help site serves these directly.
    """

    uuid = models.UUIDField(unique=True, default=uuid4)
    article = models.ForeignKey(Article, on_delete=models.PROTECT, related_name="images")
    name = models.CharField(max_length=255)
    path = models.CharField(max_length=2048)  # key in the public bucket
    content_type = models.CharField(max_length=255)
    size = models.IntegerField()

    created_by = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.PROTECT, related_name="+")
    created_on = models.DateTimeField(default=timezone.now)

    @property
    def url(self) -> str:
        return public_file_storage.url(self.path)

    def delete(self):
        path = self.path

        super().delete()

        # only remove the storage object once the row is gone
        public_file_storage.delete(path)


def get_knowledge_item_path(knowledge, item_uuid, filename: str) -> str:
    return f"orgs/{knowledge.org_id}/knowledge/{knowledge.uuid}/{item_uuid}{Path(filename).suffix.lower()}"


class KnowledgeItem(models.Model):
    """
    One page or one uploaded document in an ingested knowledge source.

    Writers are split: this app creates document rows on upload (url null, path set); mailroom creates page rows as it
    crawls (url set) and owns status/error/num_chunks for both. A page's uuid is stable across recrawls because rows
    are matched by normalised url within the source - that's what makes incremental reindexing possible, since chunks
    key off that uuid.
    """

    ALLOWED_CONTENT_TYPES = (
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # .docx
        "text/csv",
        "text/markdown",
        "text/plain",
    )
    MAX_UPLOAD_SIZE = 1024 * 1024 * 20  # 20MB
    MAX_DOCUMENTS = 100  # per documents source

    STATUS_PENDING = "P"
    STATUS_INDEXING = "I"
    STATUS_READY = "R"
    STATUS_FAILED = "F"
    STATUS_CHOICES = (
        (STATUS_PENDING, _("Pending")),
        (STATUS_INDEXING, _("Indexing")),
        (STATUS_READY, _("Ready")),
        (STATUS_FAILED, _("Failed")),
    )

    uuid = models.UUIDField(unique=True, default=uuid4)
    knowledge = models.ForeignKey(Knowledge, on_delete=models.PROTECT, related_name="items")
    name = models.CharField(max_length=255)  # page title, or the cleaned original filename

    # null for uploads; the normalised page URL for crawled pages, and their identity within the source
    url = models.URLField(max_length=2048, null=True)

    # storage key: set for uploads (private bucket); optionally the cached extracted text for pages
    path = models.CharField(max_length=2048, null=True)

    content_type = models.CharField(max_length=255)  # mailroom sets text/html for pages
    size = models.IntegerField()  # bytes of the stored/fetched content

    # written by mailroom as it extracts and chunks
    status = models.CharField(max_length=1, choices=STATUS_CHOICES, default=STATUS_PENDING)
    error = models.CharField(max_length=255, null=True)
    num_chunks = models.IntegerField(default=0)

    created_by = models.ForeignKey(  # null for crawled pages - nobody uploaded them
        settings.AUTH_USER_MODEL, on_delete=models.PROTECT, null=True, related_name="+"
    )
    created_on = models.DateTimeField(default=timezone.now)

    @classmethod
    def is_allowed_type(cls, content_type: str) -> bool:
        return content_type in cls.ALLOWED_CONTENT_TYPES

    @classmethod
    def clean_name(cls, filename: str) -> str:
        base_name, extension = os.path.splitext(filename)
        base_name = re.sub(r"[^\w\-\[\]\(\) ]", "", base_name).strip()[:200] or "file"
        return base_name + extension[:50]

    @classmethod
    def from_upload(cls, knowledge, user, file):
        assert knowledge.knowledge_type == Knowledge.TYPE_DOCUMENTS, "can only upload to documents knowledge"
        assert cls.is_allowed_type(file.content_type), "unsupported content type"

        uuid = uuid4()
        path = default_storage.save(get_knowledge_item_path(knowledge, uuid, file.name), file)

        obj = cls.objects.create(
            uuid=uuid,
            knowledge=knowledge,
            name=cls.clean_name(file.name),
            url=None,  # explicit: this is what makes it a document rather than a page
            path=path,
            content_type=file.content_type,
            size=default_storage.size(path),
            created_by=user,
        )

        knowledge.mark_pending()
        return obj

    @property
    def org(self):
        return self.knowledge.org

    def delete(self):
        path = self.path

        with transaction.atomic():
            # this item's chunks are no longer valid - mailroom will recompute the source's counters
            delete_in_batches(self.knowledge.chunks.filter(item_key=self.uuid))
            super().delete()
            self.knowledge.mark_pending()

        # only remove the storage object once the deletion has committed
        if path:
            default_storage.delete(path)

    class Meta:
        constraints = [
            # a page's identity within its source. Postgres treats NULLs as distinct in a unique index, so uploaded
            # documents (url null) are exempt automatically - no partial-index condition needed, and any number of
            # documents can coexist in one source.
            models.UniqueConstraint("knowledge", "url", name="unique_knowledge_item_urls"),
        ]


class KnowledgeChunk(models.Model):
    """
    A chunk of indexed content with its embedding. Rows are written exclusively by mailroom - this app never INSERTs
    or UPDATEs them. It only DELETEs them when the user deletes the owning item or source.
    """

    EMBEDDING_DIMENSIONS = 384  # intfloat/multilingual-e5-small

    knowledge = models.ForeignKey(Knowledge, on_delete=models.PROTECT, related_name="chunks")

    # the owning item's uuid. Not an FK, because the item lives in a different table per source type: KnowledgeItem
    # for pages/documents, Shortcut for shortcuts, Article for helpdesk. One mechanism spanning all four beats an FK
    # plus fallbacks.
    item_key = models.UUIDField()
    item_name = models.CharField(max_length=255)
    item_url = models.URLField(max_length=2048, null=True)

    text = models.TextField()
    embedding = VectorField(dimensions=EMBEDDING_DIMENSIONS)

    class Meta:
        indexes = [
            # a single ANN index shared by all orgs and sources - queries filtered by knowledge source rely on
            # pgvector >= 0.8 iterative scans (enforced by migration 0092) to return complete results
            HnswIndex(
                name="knowledgechunk_embedding",
                fields=["embedding"],
                m=16,
                ef_construction=64,
                opclasses=("vector_cosine_ops",),
            ),
            # lets mailroom replace one item's chunks on reindex, and lets us delete one item's chunks
            models.Index(name="knowledgechunk_by_item", fields=("knowledge", "item_key")),
        ]


class Topic(TembaModel, DependencyMixin):
    """
    The topic of a ticket which controls who can access that ticket.
    """

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="topics")
    is_default = models.BooleanField(default=False)

    org_limit_key = Org.LIMIT_TOPICS

    @classmethod
    def create_system(cls, org):
        assert not org.topics.filter(is_default=True).exists(), "org already has default topic"

        org.topics.create(
            name="General",
            is_default=True,
            is_system=True,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )

    @classmethod
    def create(cls, org, user, name: str):
        assert cls.is_valid_name(name), f"'{name}' is not a valid topic name"
        assert not org.topics.filter(name__iexact=name).exists(), f"topic with name '{name}' already exists"

        return org.topics.create(name=name, created_by=user, modified_by=user)

    @classmethod
    def create_from_import_def(cls, org, user, definition: dict):
        return cls.create(org, user, definition["name"])

    @classmethod
    def get_restriction(cls, org, user):
        """
        Returns the topics the given user is restricted to in the org, or None if they can access all of the org's
        topics. Staff and members whose team grants all topics are unrestricted; a member on a topic-limited team is
        restricted to that team's topics; a user with no membership in the org can access nothing. This is the single
        source of truth for team topic access, shared by everything that scopes topics or tickets to a user.
        """
        if user.is_staff:
            return None

        membership = org.get_membership(user)
        if not membership:
            return cls.objects.none()

        if membership.team and not membership.team.all_topics:
            return membership.team.topics.all()

        return None

    @classmethod
    def get_accessible(cls, org, user):
        """
        Gets the topics accessible to the given user in the given org.
        """
        restricted = cls.get_restriction(org, user)
        return org.topics.filter(is_active=True) if restricted is None else restricted

    def release(self, user):
        assert not (self.is_system and self.org.is_active), "can't release system topics"
        assert not self.tickets.exists(), "can't release topic with tickets"

        super().release(user)

        for team in self.teams.all():
            team.topics.remove(self)

        # delete ticket counts for this topic
        self.org.counts.prefix(
            [f"tickets:{Ticket.STATUS_OPEN}:{self.id}:", f"tickets:{Ticket.STATUS_CLOSED}:{self.id}:"]
        ).delete()

        self.is_active = False
        self.name = self._deleted_name()
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_by", "modified_on"))

    def as_engine_ref(self) -> dict:
        return {"uuid": str(self.uuid), "name": self.name}

    class Meta:
        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_topic_names")]


class Team(TembaModel):
    """
    Agent users are assigned to a team which controls which topics they can access.
    """

    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="teams")
    topics = models.ManyToManyField(Topic, related_name="teams")
    all_topics = models.BooleanField(default=False)
    is_default = models.BooleanField(default=False)

    org_limit_key = Org.LIMIT_TEAMS
    max_topics = 10

    @classmethod
    def create_system(cls, org):
        assert not org.teams.filter(is_default=True).exists(), "org already has default team"

        org.teams.create(
            name="All Topics",
            is_default=True,
            is_system=True,
            all_topics=True,
            created_by=org.created_by,
            modified_by=org.modified_by,
        )

    @classmethod
    def create(cls, org, user, name: str, *, topics=(), all_topics: bool = False):
        assert cls.is_valid_name(name), f"'{name}' is not a valid team name"
        assert not org.teams.filter(name__iexact=name, is_active=True).exists()
        assert not (topics and all_topics), "can't specify topics and all_topics"

        team = org.teams.create(name=name, all_topics=all_topics, created_by=user, modified_by=user)
        team.topics.add(*topics)
        return team

    def get_users(self):
        return self.org.users.filter(orgmembership__team=self)

    def release(self, user):
        assert not (self.is_system and self.org.is_active), "can't release system teams"

        # re-assign agents in this team to the default team
        OrgMembership.objects.filter(org=self.org, team=self).update(team=self.org.default_team)

        self.name = self._deleted_name()
        self.is_active = False
        self.modified_by = user
        self.save(update_fields=("name", "is_active", "modified_by", "modified_on"))

    class Meta:
        constraints = [models.UniqueConstraint("org", Lower("name"), name="unique_team_names")]


class Ticket(models.Model):
    """
    A ticket represents a period of human interaction with a contact.
    """

    STATUS_OPEN = "O"
    STATUS_CLOSED = "C"
    STATUS_CHOICES = ((STATUS_OPEN, _("Open")), (STATUS_CLOSED, _("Closed")))

    MAX_NOTE_LENGTH = 10_000

    uuid = models.UUIDField(unique=True)
    org = models.ForeignKey(Org, on_delete=models.PROTECT, related_name="tickets", db_index=False)  # indexed below
    contact = models.ForeignKey(Contact, on_delete=models.PROTECT, related_name="tickets", db_index=False)
    topic = models.ForeignKey(Topic, on_delete=models.PROTECT, related_name="tickets")

    # the status of this ticket and who it's currently assigned to
    status = models.CharField(max_length=1, choices=STATUS_CHOICES)
    assignee = models.ForeignKey(
        settings.AUTH_USER_MODEL, on_delete=models.PROTECT, null=True, related_name="assigned_tickets"
    )

    opened_on = models.DateTimeField(default=timezone.now)
    opened_in = models.ForeignKey("flows.Flow", null=True, on_delete=models.PROTECT, related_name="opened_tickets")
    opened_by = models.ForeignKey(
        settings.AUTH_USER_MODEL, null=True, on_delete=models.PROTECT, related_name="opened_tickets"
    )

    # when this ticket was first replied to, closed, modified
    replied_on = models.DateTimeField(null=True)
    closed_on = models.DateTimeField(null=True)
    modified_on = models.DateTimeField(default=timezone.now)

    # when this ticket last had activity which includes messages being sent and received, and is used for ordering
    last_activity_on = models.DateTimeField(default=timezone.now)

    def add_note(self, user: User, *, note: str):
        self.bulk_add_note(self.org, user, [self], note=note)

    @classmethod
    def bulk_assign(cls, org, user: User, tickets: list, assignee: User, *, via_api=False) -> list[str]:
        return cls._bulk_response(
            mailroom.get_client().ticket_change_assignee(org, user, tickets, assignee, via="api" if via_api else "ui")
        )

    @classmethod
    def bulk_add_note(cls, org, user: User, tickets: list, note: str, *, via_api=False) -> list[str]:
        return cls._bulk_response(
            mailroom.get_client().ticket_add_note(org, user, tickets, note, via="api" if via_api else "ui")
        )

    @classmethod
    def bulk_change_topic(cls, org, user: User, tickets: list, topic: Topic, *, via_api=False) -> list[str]:
        return cls._bulk_response(
            mailroom.get_client().ticket_change_topic(org, user, tickets, topic, via="api" if via_api else "ui")
        )

    @classmethod
    def bulk_close(cls, org, user, tickets, *, via_api=False) -> list[str]:
        return cls._bulk_response(
            mailroom.get_client().ticket_close(org, user, tickets, via="api" if via_api else "ui")
        )

    @classmethod
    def bulk_reopen(cls, org, user, tickets, *, via_api=False) -> list[str]:
        return cls._bulk_response(
            mailroom.get_client().ticket_reopen(org, user, tickets, via="api" if via_api else "ui")
        )

    @classmethod
    def _bulk_response(cls, resp: dict) -> list[str]:
        return resp.get("changed_uuids", [])

    @classmethod
    def get_accessible(cls, org, user):
        """
        Gets the tickets in the org that the given user is allowed to view. This mirrors what the ticketing UI exposes:
        the union of the Mine and All folders. Staff and users whose team grants all topics can view every ticket;
        an agent on a topic-restricted team can view tickets in their team's topics plus any assigned to them (they can
        always see their own tickets even in topics they otherwise lack access to). A user with no membership in the org
        can view nothing - we fail closed rather than exposing the whole workspace.
        """
        qs = org.tickets.all()

        restricted = Topic.get_restriction(org, user)
        if restricted is not None:
            qs = qs.filter(Q(assignee=user) | Q(topic__in=restricted))

        return qs

    @classmethod
    def get_assignee_count(cls, org, user, topics, status: str) -> int:
        """
        Gets the count of tickets assigned to the given user across the given topics and status.
        """
        return org.counts.filter(scope__in=[f"tickets:{status}:{t.id}:{user.id if user else 0}" for t in topics]).sum()

    @classmethod
    def get_status_count(cls, org, topics, status: str) -> int:
        """
        Gets the count across the given topics and status.
        """
        return org.counts.prefix([f"tickets:{status}:{t.id}:" for t in topics]).sum()

    @classmethod
    def get_topic_counts(cls, org, topics, status: str) -> dict[Topic, int]:
        """
        Gets the count for each topic and the given status.
        """

        # count scopes are stored as 'tickets:<status>:<topic-id>:<assignee-id>' so get all counts with the prefix
        # 'tickets:<status>:' and group by topic-id extracted as the 3rd split part.
        counts = (
            org.counts.prefix(f"tickets:{status}:")
            .annotate(topic_id=Cast(SplitPart(F("scope"), Value(":"), Value(3)), output_field=models.IntegerField()))
            .values_list("topic_id")
            .annotate(count_sum=Sum("count"))
        )
        by_topic_id = {c[0]: c[1] for c in counts}
        return {t: by_topic_id.get(t.id, 0) for t in topics}

    def __str__(self):
        return f"Ticket[uuid={self.uuid}, topic={self.topic.name}]"

    class Meta:
        indexes = [
            # used by the All folder - status is descending so that a forward scan yields the ticket UI's
            # display order of open ('O') before closed ('C'), then most recent activity.
            #
            # both the column order and the descending directions of these two indexes are load-bearing for the
            # folder view's keyset cursor (the RawSQL row comparison in tickets/views.py) - it relies on a forward
            # scan of the index already being in display order, so changing either the order of the columns or
            # their directions silently degrades folder paging to sort-based plans.
            models.Index(name="tickets_org_status_desc", fields=["org", "-status", "-last_activity_on", "-id"]),
            # used by the Unassigned and Mine folders
            models.Index(
                name="tickets_org_assign_status_desc",
                fields=["org", "assignee", "-status", "-last_activity_on", "-id"],
            ),
            # used by engine to load a contact with its open tickets
            models.Index(name="tickets_contact_open", fields=["contact", "opened_on"], condition=Q(status="O")),
            # used by API tickets endpoint hence the ordering, and general fetching by org or contact
            models.Index(name="tickets_api_by_org", fields=["org", "-modified_on", "-id"]),
            models.Index(name="tickets_api_by_contact", fields=["contact", "-modified_on", "-id"]),
        ]


class TicketFolder(metaclass=ABCMeta):
    slug = None
    name = None
    icon = None
    verbose_name = None
    restrict_topics = None

    def get_icon(self, count) -> str:
        return self.icon

    def get_queryset(self, org, user, *, ordered: bool):
        qs = org.tickets.all()

        if self.restrict_topics:
            restricted = Topic.get_restriction(org, user)
            if restricted is not None:
                qs = qs.filter(topic__in=restricted)

        if ordered:
            # the ticket UI's display order: open ('O' > 'C') before closed, then most recent activity - for
            # status filtered fetches the leading term is constant so this is just most recent activity
            qs = qs.order_by("-status", "-last_activity_on", "-id")

        return qs.select_related("topic", "assignee").prefetch_related("contact")

    @classmethod
    def from_slug(cls, org, user, slug_or_uuid: str):
        if is_uuid(slug_or_uuid):
            topic = Topic.get_accessible(org, user).filter(uuid=slug_or_uuid).first()
            if topic:
                return TopicFolder(topic)

        return FOLDERS.get(slug_or_uuid, None)

    @classmethod
    def all(cls):
        return FOLDERS


class MineFolder(TicketFolder):
    """
    Tickets assigned to the current user.
    """

    slug = "mine"
    name = _("My Tickets")
    icon = "tickets_mine"
    restrict_topics = False  # users can see tickets assigned to them even if they don't have access to the topic

    def get_icon(self, count) -> str:
        return self.icon if count else "tickets_mine_done"

    def get_queryset(self, org, user, *, ordered: bool):
        return super().get_queryset(org, user, ordered=ordered).filter(assignee=user)


class UnassignedFolder(TicketFolder):
    """
    Tickets not assigned to any user.
    """

    slug = "unassigned"
    name = _("Unassigned")
    verbose_name = _("Unassigned Tickets")
    icon = "tickets_unassigned"
    restrict_topics = True

    def get_queryset(self, org, user, *, ordered: bool):
        return super().get_queryset(org, user, ordered=ordered).filter(assignee=None)


class AllFolder(TicketFolder):
    """
    All tickets the user can access.
    """

    slug = "all"
    name = _("All")
    verbose_name = _("All Tickets")
    icon = "tickets_all"
    restrict_topics = True


FOLDERS = {f.slug: f() for f in TicketFolder.__subclasses__()}


class TopicFolder(TicketFolder):
    """
    Wraps a topic so we can use it like a folder.
    """

    def __init__(self, topic: Topic):
        self.slug = topic.uuid
        self.name = topic.name
        self.topic = topic
        self.restrict_topics = False  # already filtered by a single topic

    def get_queryset(self, org, user, *, ordered: bool):
        return super().get_queryset(org, user, ordered=ordered).filter(topic=self.topic)


def export_ticket_stats(org: Org, since: date, until: date) -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tickets"
    sheet.merge_cells("A1:A2")
    sheet.cell(row=1, column=2, value="Workspace")
    sheet.merge_cells("B1:D1")
    sheet.cell(row=2, column=2, value="Opened")
    sheet.cell(row=2, column=3, value="Replies")
    sheet.cell(row=2, column=4, value="Reply Time (Secs)")

    users = list(org.users.order_by("email"))

    user_col = 5
    for user in users:
        cell = sheet.cell(row=1, column=user_col, value=str(user))
        cell.hyperlink = f"mailto:{user.email}"
        cell.style = "Hyperlink"
        sheet.merge_cells(start_row=1, start_column=user_col, end_row=1, end_column=user_col + 1)

        sheet.cell(row=2, column=user_col, value="Assigned")
        sheet.cell(row=2, column=user_col + 1, value="Replies")
        user_col += 2

    org_openings = org.daily_counts.period(since, until).prefix("tickets:opened:").day_totals(scoped=False)
    all_replies = org.daily_counts.period(since, until).prefix("msgs:ticketreplies:").day_totals(scoped=True)
    all_assignments = org.daily_counts.period(since, until).prefix("tickets:assigned:").day_totals(scoped=True)
    all_resptimes = org.daily_counts.period(since, until).prefix("ticketresptime:").day_totals(scoped=True)

    user_assignments = defaultdict(dict)
    for (day, scope), count in all_assignments.items():
        user_id = int(scope.split(":")[-1])
        user_assignments[user_id][day] = count

    org_replies = defaultdict(int)
    user_replies = defaultdict(dict)
    for (day, scope), count in all_replies.items():
        user_id = int(scope.split(":")[-1])
        user_replies[user_id][day] = count
        org_replies[day] += count

    org_resptimes, org_respcounts = defaultdict(int), defaultdict(int)
    for (day, scope), count in all_resptimes.items():
        if scope.endswith(":total"):
            org_resptimes[day] += count
        elif scope.endswith(":count"):
            org_respcounts[day] += count

    org_respavgs = {}
    for day, total in org_resptimes.items():
        org_respavgs[day] = total // org_respcounts[day]

    day_row = 3
    for day in date_range(since, until):
        sheet.cell(row=day_row, column=1, value=day)
        sheet.cell(row=day_row, column=2, value=org_openings.get(day, 0))
        sheet.cell(row=day_row, column=3, value=org_replies.get(day, 0))
        sheet.cell(row=day_row, column=4, value=org_respavgs.get(day, ""))

        user_col = 5
        for user in users:
            sheet.cell(row=day_row, column=user_col, value=user_assignments[user.id].get(day, 0))
            sheet.cell(row=day_row, column=user_col + 1, value=user_replies[user.id].get(day, 0))
            user_col += 2

        day_row += 1

    return workbook


class TicketExport(ExportType):
    """
    Export of tickets
    """

    slug = "ticket"
    name = _("Tickets")
    download_prefix = "tickets"

    @classmethod
    def create(cls, org, user, start_date, end_date, with_fields=(), with_groups=()):
        return Export.objects.create(
            org=org,
            export_type=cls.slug,
            start_date=start_date,
            end_date=end_date,
            config={"with_fields": [f.id for f in with_fields], "with_groups": [g.id for g in with_groups]},
            created_by=user,
        )

    def write(self, export):
        headers = ["UUID", "Opened On", "Closed On", "Topic", "Assigned To"] + export.get_contact_headers()
        start_date, end_date = export.get_date_range()

        # get the ticket ids, filtered and ordered by opened on
        ticket_ids = (
            Ticket.objects.filter(org=export.org, opened_on__gte=start_date, opened_on__lte=end_date)
            .order_by("opened_on")
            .values_list("id", flat=True)
            .using("readonly")
        )

        exporter = MultiSheetExporter("Tickets", headers, export.org.timezone)
        num_records = 0

        # add tickets to the export in batches of 1k to limit memory usage
        for batch_ids in itertools.batched(ticket_ids, 1000):
            tickets = (
                Ticket.objects.filter(id__in=batch_ids)
                .order_by("opened_on")
                .prefetch_related("org", "contact", "contact__org", "contact__groups", "assignee", "topic")
                .using("readonly")
            )

            Contact.bulk_urn_cache_initialize([t.contact for t in tickets], using="readonly")

            for ticket in tickets:
                values = [
                    str(ticket.uuid),
                    ticket.opened_on,
                    ticket.closed_on,
                    ticket.topic.name,
                    ticket.assignee.email if ticket.assignee else None,
                ]
                values += export.get_contact_columns(ticket.contact)

                exporter.write_row(values)

            num_records += len(tickets)

            export.modified_on = timezone.now()
            export.save(update_fields=("modified_on",))

        return *exporter.save_file(), num_records
