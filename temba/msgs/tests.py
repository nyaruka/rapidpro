import random
import string
from datetime import date, datetime, timedelta
from unittest.mock import patch

import pytz
from openpyxl import load_workbook

from django.conf import settings
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone

from temba.archives.models import Archive
from temba.channels.models import ChannelCount, ChannelEvent, ChannelLog
from temba.contacts.models import URN, Contact, ContactURN
from temba.contacts.search.omnibox import omnibox_serialize
from temba.flows.models import Flow
from temba.msgs.models import (
    Attachment,
    Broadcast,
    ExportMessagesTask,
    Label,
    LabelCount,
    Media,
    Msg,
    SystemLabel,
    SystemLabelCount,
)
from temba.schedules.models import Schedule
from temba.tests import AnonymousOrg, CRUDLTestMixin, TembaTest, mock_uuids
from temba.tests.engine import MockSessionWriter
from temba.tests.s3 import MockS3Client, jsonlgz_encode
from temba.tickets.models import Ticket
from temba.utils import s3
from temba.utils.compose import compose_deserialize_attachments, compose_serialize
from temba.utils.views import TEMBA_MENU_SELECTION

from .tasks import fail_old_messages, squash_msg_counts
from .templatetags.sms import as_icon


class AttachmentTest(TembaTest):
    def test_attachments(self):
        # check equality
        self.assertEqual(
            Attachment("image/jpeg", "http://example.com/test.jpg"),
            Attachment("image/jpeg", "http://example.com/test.jpg"),
        )

        # check parsing
        self.assertEqual(
            Attachment("image", "http://example.com/test.jpg"),
            Attachment.parse("image:http://example.com/test.jpg"),
        )
        self.assertEqual(
            Attachment("image/jpeg", "http://example.com/test.jpg"),
            Attachment.parse("image/jpeg:http://example.com/test.jpg"),
        )
        with self.assertRaises(ValueError):
            Attachment.parse("http://example.com/test.jpg")


class MediaTest(TembaTest):
    def tearDown(self):
        self.clear_storage()

        return super().tearDown()

    def test_clean_name(self):
        self.assertEqual("file.jpg", Media.clean_name("", "image/jpeg"))
        self.assertEqual("foo.jpg", Media.clean_name("foo", "image/jpeg"))
        self.assertEqual("file.png", Media.clean_name("*.png", "image/png"))
        self.assertEqual("passwd.jpg", Media.clean_name(".passwd", "image/jpeg"))
        self.assertEqual("tést[0].jpg", Media.clean_name("tést[0]/^..\\", "image/jpeg"))

    @mock_uuids
    def test_from_upload(self):
        media = Media.from_upload(
            self.org,
            self.admin,
            self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
            process=False,
        )

        self.assertEqual("b97f69f7-5edf-45c7-9fda-d37066eae91d", str(media.uuid))
        self.assertEqual(self.org, media.org)
        self.assertEqual(
            f"/media/test_orgs/{self.org.id}/media/b97f/b97f69f7-5edf-45c7-9fda-d37066eae91d/steve%20marten.jpg",
            media.url,
        )
        self.assertEqual("image/jpeg", media.content_type)
        self.assertEqual(
            f"test_orgs/{self.org.id}/media/b97f/b97f69f7-5edf-45c7-9fda-d37066eae91d/steve marten.jpg", media.path
        )
        self.assertEqual(self.admin, media.created_by)
        self.assertEqual(Media.STATUS_PENDING, media.status)

        # check that our filename is cleaned
        media = Media.from_upload(
            self.org,
            self.admin,
            self.upload(f"{settings.MEDIA_ROOT}/test_media/klab.png", "image/png", name="../../../etc/passwd"),
            process=False,
        )

        self.assertEqual(
            f"test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/passwd.png", media.path
        )

    @mock_uuids
    def test_process_image_png(self):
        media = Media.from_upload(
            self.org,
            self.admin,
            self.upload(f"{settings.MEDIA_ROOT}/test_media/klab.png", "image/png"),
        )
        media.refresh_from_db()

        self.assertEqual(371425, media.size)
        self.assertEqual(0, media.duration)
        self.assertEqual(480, media.width)
        self.assertEqual(360, media.height)
        self.assertEqual(Media.STATUS_READY, media.status)

    @mock_uuids
    def test_process_audio_wav(self):
        media = Media.from_upload(
            self.org, self.admin, self.upload(f"{settings.MEDIA_ROOT}/test_media/allo.wav", "audio/wav")
        )
        media.refresh_from_db()

        self.assertEqual(81818, media.size)
        self.assertEqual(5110, media.duration)
        self.assertEqual(0, media.width)
        self.assertEqual(0, media.height)
        self.assertEqual(Media.STATUS_READY, media.status)

        alt1, alt2 = list(media.alternates.order_by("id"))

        self.assertEqual(self.org, alt1.org)
        self.assertEqual(
            f"/media/test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/allo.mp3", alt1.url
        )
        self.assertEqual("audio/mp3", alt1.content_type)
        self.assertEqual(
            f"test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/allo.mp3", alt1.path
        )
        self.assertAlmostEqual(5517, alt1.size, delta=1000)
        self.assertEqual(5110, alt1.duration)
        self.assertEqual(0, alt1.width)
        self.assertEqual(0, alt1.height)
        self.assertEqual(Media.STATUS_READY, alt1.status)

        self.assertEqual(self.org, alt2.org)
        self.assertEqual(
            f"/media/test_orgs/{self.org.id}/media/d1ee/d1ee73f0-bdb5-47ce-99dd-0c95d4ebf008/allo.m4a", alt2.url
        )
        self.assertEqual("audio/mp4", alt2.content_type)
        self.assertEqual(
            f"test_orgs/{self.org.id}/media/d1ee/d1ee73f0-bdb5-47ce-99dd-0c95d4ebf008/allo.m4a", alt2.path
        )
        self.assertAlmostEqual(20552, alt2.size, delta=7500)
        self.assertEqual(5110, alt2.duration)
        self.assertEqual(0, alt2.width)
        self.assertEqual(0, alt2.height)
        self.assertEqual(Media.STATUS_READY, alt2.status)

    @mock_uuids
    def test_process_audio_m4a(self):
        media = Media.from_upload(
            self.org, self.admin, self.upload(f"{settings.MEDIA_ROOT}/test_media/bubbles.m4a", "audio/mp4")
        )
        media.refresh_from_db()

        self.assertEqual(46468, media.size)
        self.assertEqual(10216, media.duration)
        self.assertEqual(0, media.width)
        self.assertEqual(0, media.height)
        self.assertEqual(Media.STATUS_READY, media.status)

        alt = media.alternates.get()

        self.assertEqual(self.org, alt.org)
        self.assertEqual(
            f"/media/test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/bubbles.mp3", alt.url
        )
        self.assertEqual("audio/mp3", alt.content_type)
        self.assertEqual(
            f"test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/bubbles.mp3", alt.path
        )
        self.assertAlmostEqual(41493, alt.size, delta=1000)
        self.assertEqual(10216, alt.duration)
        self.assertEqual(0, alt.width)
        self.assertEqual(0, alt.height)
        self.assertEqual(Media.STATUS_READY, alt.status)

    @mock_uuids
    def test_process_video_mp4(self):
        media = Media.from_upload(
            self.org, self.admin, self.upload(f"{settings.MEDIA_ROOT}/test_media/snow.mp4", "video/mp4")
        )
        media.refresh_from_db()

        self.assertEqual(684558, media.size)
        self.assertEqual(3536, media.duration)
        self.assertEqual(640, media.width)
        self.assertEqual(480, media.height)
        self.assertEqual(Media.STATUS_READY, media.status)

        alt = media.alternates.get()

        self.assertEqual(self.org, alt.org)
        self.assertEqual(
            f"/media/test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/snow.jpg", alt.url
        )
        self.assertEqual("image/jpeg", alt.content_type)
        self.assertEqual(f"test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/snow.jpg", alt.path)
        self.assertAlmostEqual(37613, alt.size, delta=1000)
        self.assertEqual(0, alt.duration)
        self.assertEqual(640, alt.width)
        self.assertEqual(480, alt.height)
        self.assertEqual(Media.STATUS_READY, alt.status)

    @mock_uuids
    def test_process_unsupported(self):
        media = Media.from_upload(
            self.org, self.admin, self.upload(f"{settings.MEDIA_ROOT}/test_imports/simple.xls", "audio/m4a")
        )
        media.refresh_from_db()

        self.assertEqual(19968, media.size)
        self.assertEqual(Media.STATUS_FAILED, media.status)


class MsgTest(TembaTest, CRUDLTestMixin):
    def setUp(self):
        super().setUp()

        self.joe = self.create_contact("Joe Blow", phone="123")
        ContactURN.create(self.org, self.joe, "tel:789")

        self.frank = self.create_contact("Frank Blow", phone="321")
        self.kevin = self.create_contact("Kevin Durant", phone="987")

        self.just_joe = self.create_group("Just Joe", [self.joe])
        self.joe_and_frank = self.create_group("Joe and Frank", [self.joe, self.frank])

    def test_as_archive_json(self):
        flow = self.create_flow("Color Flow")
        msg1 = self.create_incoming_msg(self.joe, "i'm having a problem", flow=flow)
        self.assertEqual(
            {
                "id": msg1.id,
                "contact": {"uuid": str(self.joe.uuid), "name": "Joe Blow"},
                "channel": {"uuid": str(self.channel.uuid), "name": "Test Channel"},
                "flow": {"uuid": str(flow.uuid), "name": "Color Flow"},
                "urn": "tel:123",
                "direction": "in",
                "type": "text",
                "status": "handled",
                "visibility": "visible",
                "text": "i'm having a problem",
                "attachments": [],
                "labels": [],
                "created_on": msg1.created_on.isoformat(),
                "sent_on": None,
            },
            msg1.as_archive_json(),
        )

        # label first message
        label = self.create_label("la\02bel1")
        label.toggle_label([msg1], add=True)

        self.assertEqual(
            {
                "id": msg1.id,
                "contact": {"uuid": str(self.joe.uuid), "name": "Joe Blow"},
                "channel": {"uuid": str(self.channel.uuid), "name": "Test Channel"},
                "flow": {"uuid": str(flow.uuid), "name": "Color Flow"},
                "urn": "tel:123",
                "direction": "in",
                "type": "text",
                "status": "handled",
                "visibility": "visible",
                "text": "i'm having a problem",
                "attachments": [],
                "labels": [{"uuid": str(label.uuid), "name": "la\x02bel1"}],
                "created_on": msg1.created_on.isoformat(),
                "sent_on": None,
            },
            msg1.as_archive_json(),
        )

        msg2 = self.create_incoming_msg(
            self.joe, "Media message", attachments=["audio:http://rapidpro.io/audio/sound.mp3"]
        )

        self.assertEqual(
            {
                "id": msg2.id,
                "contact": {"uuid": str(self.joe.uuid), "name": "Joe Blow"},
                "channel": {"uuid": str(self.channel.uuid), "name": "Test Channel"},
                "flow": None,
                "urn": "tel:123",
                "direction": "in",
                "type": "text",
                "status": "handled",
                "visibility": "visible",
                "text": "Media message",
                "attachments": [{"url": "http://rapidpro.io/audio/sound.mp3", "content_type": "audio"}],
                "labels": [],
                "created_on": msg2.created_on.isoformat(),
                "sent_on": None,
            },
            msg2.as_archive_json(),
        )

    @patch("django.core.files.storage.default_storage.delete")
    def test_bulk_soft_delete(self, mock_storage_delete):
        # create some messages
        msg1 = self.create_incoming_msg(
            self.joe,
            "i'm having a problem",
            attachments=[
                r"audo/mp4:http://s3.com/attachments/1/a/b.jpg",
                r"image/jpeg:http://s3.com/attachments/1/c/d%20e.jpg",
            ],
        )
        msg2 = self.create_incoming_msg(self.frank, "ignore joe, he's a liar")
        out1 = self.create_outgoing_msg(self.frank, "hi")

        # can't soft delete outgoing messages
        with self.assertRaises(AssertionError):
            Msg.bulk_soft_delete([out1])

        Msg.bulk_soft_delete([msg1, msg2])

        # soft delete should clear text and attachments
        for msg in (msg1, msg2):
            msg.refresh_from_db()

            self.assertEqual("", msg.text)
            self.assertEqual([], msg.attachments)
            self.assertEqual(Msg.VISIBILITY_DELETED_BY_USER, msg1.visibility)

        mock_storage_delete.assert_any_call("/attachments/1/a/b.jpg")
        mock_storage_delete.assert_any_call("/attachments/1/c/d e.jpg")

    @patch("django.core.files.storage.default_storage.delete")
    def test_bulk_delete(self, mock_storage_delete):
        # create some messages
        msg1 = self.create_incoming_msg(
            self.joe,
            "i'm having a problem",
            attachments=[
                r"audo/mp4:http://s3.com/attachments/1/a/b.jpg",
                r"image/jpeg:http://s3.com/attachments/1/c/d%20e.jpg",
            ],
        )
        msg2 = self.create_incoming_msg(self.frank, "ignore joe, he's a liar")
        out1 = self.create_outgoing_msg(self.frank, "hi")

        # create a channel log for each message
        ChannelLog.objects.create(channel=self.channel, msg=msg1, is_error=False)
        ChannelLog.objects.create(channel=self.channel, msg=msg2, is_error=False)
        ChannelLog.objects.create(channel=self.channel, msg=out1, is_error=False)

        Msg.bulk_delete([msg1, out1])

        self.assertEqual(1, Msg.objects.all().count())
        self.assertEqual(1, ChannelLog.objects.filter(msg_id=msg2.id).count())

        mock_storage_delete.assert_any_call("/attachments/1/a/b.jpg")
        mock_storage_delete.assert_any_call("/attachments/1/c/d e.jpg")

    def test_archive_and_release(self):
        msg1 = self.create_incoming_msg(self.joe, "Incoming")
        label = self.create_label("Spam")
        label.toggle_label([msg1], add=True)

        msg1.archive()

        msg1 = Msg.objects.get(pk=msg1.pk)
        self.assertEqual(msg1.visibility, Msg.VISIBILITY_ARCHIVED)
        self.assertEqual(set(msg1.labels.all()), {label})  # don't remove labels

        msg1.restore()

        msg1 = Msg.objects.get(pk=msg1.id)
        self.assertEqual(msg1.visibility, Msg.VISIBILITY_VISIBLE)

        msg1.delete()
        self.assertFalse(Msg.objects.filter(pk=msg1.pk).exists())

        label.refresh_from_db()
        self.assertEqual(0, label.get_messages().count())  # do remove labels
        self.assertIsNotNone(label)

        # can't archive outgoing messages
        msg2 = self.create_outgoing_msg(self.joe, "Outgoing")
        self.assertRaises(AssertionError, msg2.archive)

    def test_release_counts(self):
        flow = self.create_flow("Test")

        def assertReleaseCount(direction, status, visibility, flow, label):
            if direction == Msg.DIRECTION_OUT:
                msg = self.create_outgoing_msg(self.joe, "Whattup Joe", flow=flow, status=status)
            else:
                msg = self.create_incoming_msg(self.joe, "Hey hey", flow=flow, status=status)

            Msg.objects.filter(id=msg.id).update(visibility=visibility)

            # assert our folder count is right
            counts = SystemLabel.get_counts(self.org)
            self.assertEqual(counts[label], 1)

            # delete the msg, count should now be 0
            msg.delete()
            counts = SystemLabel.get_counts(self.org)
            self.assertEqual(counts[label], 0)

        # outgoing labels
        assertReleaseCount("O", Msg.STATUS_SENT, Msg.VISIBILITY_VISIBLE, None, SystemLabel.TYPE_SENT)
        assertReleaseCount("O", Msg.STATUS_QUEUED, Msg.VISIBILITY_VISIBLE, None, SystemLabel.TYPE_OUTBOX)
        assertReleaseCount("O", Msg.STATUS_FAILED, Msg.VISIBILITY_VISIBLE, flow, SystemLabel.TYPE_FAILED)

        # incoming labels
        assertReleaseCount("I", Msg.STATUS_HANDLED, Msg.VISIBILITY_VISIBLE, None, SystemLabel.TYPE_INBOX)
        assertReleaseCount("I", Msg.STATUS_HANDLED, Msg.VISIBILITY_ARCHIVED, None, SystemLabel.TYPE_ARCHIVED)
        assertReleaseCount("I", Msg.STATUS_HANDLED, Msg.VISIBILITY_VISIBLE, flow, SystemLabel.TYPE_FLOWS)

    @patch("temba.utils.email.send_temba_email")
    def test_message_export_from_archives(self, mock_send_temba_email):
        export_url = reverse("msgs.msg_export")

        self.login(self.admin)

        self.joe.name = "Jo\02e Blow"
        self.joe.save(update_fields=("name",))

        self.org.created_on = datetime(2017, 1, 1, 9, tzinfo=pytz.UTC)
        self.org.save()

        flow = self.create_flow("Color Flow")

        msg1 = self.create_incoming_msg(self.joe, "hello 1", created_on=datetime(2017, 1, 1, 10, tzinfo=pytz.UTC))
        msg2 = self.create_incoming_msg(
            self.frank, "hello 2", created_on=datetime(2017, 1, 2, 10, tzinfo=pytz.UTC), flow=flow
        )
        msg3 = self.create_incoming_msg(self.joe, "hello 3", created_on=datetime(2017, 1, 3, 10, tzinfo=pytz.UTC))

        # inbound message that looks like a surveyor message
        msg4 = self.create_incoming_msg(
            self.joe, "hello 4", surveyor=True, created_on=datetime(2017, 1, 4, 10, tzinfo=pytz.UTC)
        )

        # inbound message with media attached, such as an ivr recording
        msg5 = self.create_incoming_msg(
            self.joe,
            "Media message",
            attachments=["audio:http://rapidpro.io/audio/sound.mp3"],
            created_on=datetime(2017, 1, 5, 10, tzinfo=pytz.UTC),
        )

        # create some outbound messages with different statuses
        msg6 = self.create_outgoing_msg(
            self.joe, "Hey out 6", status=Msg.STATUS_SENT, created_on=datetime(2017, 1, 6, 10, tzinfo=pytz.UTC)
        )
        msg7 = self.create_outgoing_msg(
            self.joe, "Hey out 7", status=Msg.STATUS_DELIVERED, created_on=datetime(2017, 1, 7, 10, tzinfo=pytz.UTC)
        )
        msg8 = self.create_outgoing_msg(
            self.joe, "Hey out 8", status=Msg.STATUS_ERRORED, created_on=datetime(2017, 1, 8, 10, tzinfo=pytz.UTC)
        )
        msg9 = self.create_outgoing_msg(
            self.joe, "Hey out 9", status=Msg.STATUS_FAILED, created_on=datetime(2017, 1, 9, 10, tzinfo=pytz.UTC)
        )

        self.assertEqual(msg5.get_attachments(), [Attachment("audio", "http://rapidpro.io/audio/sound.mp3")])

        # label first message
        label = self.create_label("la\02bel1")
        label.toggle_label([msg1], add=True)

        # archive last message
        msg3.visibility = Msg.VISIBILITY_ARCHIVED
        msg3.save()

        # archive 5 msgs
        mock_s3 = MockS3Client()
        body, md5, size = jsonlgz_encode([m.as_archive_json() for m in (msg1, msg2, msg3, msg4, msg5, msg6)])
        mock_s3.put_object("test-bucket", "archive1.jsonl.gz", body)

        Archive.objects.create(
            org=self.org,
            archive_type=Archive.TYPE_MSG,
            size=size,
            hash=md5,
            url="http://test-bucket.aws.com/archive1.jsonl.gz",
            record_count=6,
            start_date=msg5.created_on.date(),
            period="D",
            build_time=23425,
        )

        with patch("django.core.files.storage.default_storage.delete"):
            msg2.delete()
            msg3.delete()
            msg4.delete()
            msg5.delete()
            msg6.delete()

        # create an archive earlier than our flow created date so we check that it isn't included
        body, md5, size = jsonlgz_encode([msg7.as_archive_json()])
        Archive.objects.create(
            org=self.org,
            archive_type=Archive.TYPE_MSG,
            size=size,
            hash=md5,
            url="http://test-bucket.aws.com/archive2.jsonl.gz",
            record_count=1,
            start_date=self.org.created_on - timedelta(days=2),
            period="D",
            build_time=5678,
        )
        mock_s3.put_object("test-bucket", "archive2.jsonl.gz", body)

        msg7.delete()

        def request_export(query, data=None):
            with self.mockReadOnly():
                response = self.client.post(export_url + query, data)
            self.assertModalResponse(response, redirect="/msg/")
            task = ExportMessagesTask.objects.order_by("-id").first()
            filename = "%s/test_orgs/%d/message_exports/%s.xlsx" % (settings.MEDIA_ROOT, self.org.id, task.uuid)
            return load_workbook(filename=filename)

        # export all visible messages (i.e. not msg3) using export_all param
        with self.assertNumQueries(34):
            with patch("temba.utils.s3.client", return_value=mock_s3):
                workbook = request_export(
                    "?l=I", {"export_all": 1, "start_date": "2000-09-01", "end_date": "2022-09-01"}
                )

        expected_headers = [
            "Date",
            "Contact UUID",
            "Contact Name",
            "URN Scheme",
            "URN Value",
            "Flow",
            "Direction",
            "Text",
            "Attachments",
            "Status",
            "Channel",
            "Labels",
        ]

        self.assertExcelSheet(
            workbook.worksheets[0],
            [
                expected_headers,
                [
                    msg1.created_on,
                    msg1.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "IN",
                    "hello 1",
                    "",
                    "handled",
                    "Test Channel",
                    "label1",
                ],
                [
                    msg2.created_on,
                    msg2.contact.uuid,
                    "Frank Blow",
                    "tel",
                    "321",
                    "Color Flow",
                    "IN",
                    "hello 2",
                    "",
                    "handled",
                    "Test Channel",
                    "",
                ],
                [msg4.created_on, msg1.contact.uuid, "Joe Blow", "", "", "", "IN", "hello 4", "", "handled", "", ""],
                [
                    msg5.created_on,
                    msg5.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "IN",
                    "Media message",
                    "http://rapidpro.io/audio/sound.mp3",
                    "handled",
                    "Test Channel",
                    "",
                ],
                [
                    msg6.created_on,
                    msg6.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "OUT",
                    "Hey out 6",
                    "",
                    "sent",
                    "Test Channel",
                    "",
                ],
                [
                    msg8.created_on,
                    msg8.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "OUT",
                    "Hey out 8",
                    "",
                    "errored",
                    "Test Channel",
                    "",
                ],
                [
                    msg9.created_on,
                    msg9.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "OUT",
                    "Hey out 9",
                    "",
                    "failed",
                    "Test Channel",
                    "",
                ],
            ],
            self.org.timezone,
        )

        with patch("temba.utils.s3.client", return_value=mock_s3):
            workbook = request_export(
                "?l=I",
                {
                    "export_all": 0,
                    "start_date": msg5.created_on.strftime("%Y-%m-%d"),
                    "end_date": msg7.created_on.strftime("%Y-%m-%d"),
                },
            )

        self.assertExcelSheet(
            workbook.worksheets[0],
            [
                expected_headers,
                [
                    msg5.created_on,
                    msg5.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "IN",
                    "Media message",
                    "http://rapidpro.io/audio/sound.mp3",
                    "handled",
                    "Test Channel",
                    "",
                ],
            ],
            self.org.timezone,
        )

        with patch("temba.utils.s3.client", return_value=mock_s3):
            workbook = request_export("?l=S", {"export_all": 0, "start_date": "2000-09-01", "end_date": "2022-09-01"})

        self.assertExcelSheet(
            workbook.worksheets[0],
            [
                expected_headers,
                [
                    msg6.created_on,
                    msg6.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "OUT",
                    "Hey out 6",
                    "",
                    "sent",
                    "Test Channel",
                    "",
                ],
            ],
            self.org.timezone,
        )

        with patch("temba.utils.s3.client", return_value=mock_s3):
            workbook = request_export("?l=X", {"export_all": 0, "start_date": "2000-09-01", "end_date": "2022-09-01"})

        self.assertExcelSheet(
            workbook.worksheets[0],
            [
                expected_headers,
                [
                    msg9.created_on,
                    msg9.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "OUT",
                    "Hey out 9",
                    "",
                    "failed",
                    "Test Channel",
                    "",
                ],
            ],
            self.org.timezone,
        )

        with patch("temba.utils.s3.client", return_value=mock_s3):
            workbook = request_export("?l=W", {"export_all": 0, "start_date": "2000-09-01", "end_date": "2022-09-01"})

        self.assertExcelSheet(
            workbook.worksheets[0],
            [
                expected_headers,
                [
                    msg2.created_on,
                    msg2.contact.uuid,
                    "Frank Blow",
                    "tel",
                    "321",
                    "Color Flow",
                    "IN",
                    "hello 2",
                    "",
                    "handled",
                    "Test Channel",
                    "",
                ],
            ],
            self.org.timezone,
        )

        with patch("temba.utils.s3.client", return_value=mock_s3):
            workbook = request_export(
                f"?l={label.uuid}", {"export_all": 0, "start_date": "2000-09-01", "end_date": "2022-09-01"}
            )

        self.assertExcelSheet(
            workbook.worksheets[0],
            [
                expected_headers,
                [
                    msg1.created_on,
                    msg1.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    "IN",
                    "hello 1",
                    "",
                    "handled",
                    "Test Channel",
                    "label1",
                ],
            ],
            self.org.timezone,
        )

        self.clear_storage()

    @patch("temba.utils.email.send_temba_email")
    def test_message_export(self, mock_send_temba_email):
        export_url = reverse("msgs.msg_export")

        self.login(self.admin)

        age = self.create_field("age", "Age")
        bob = self.create_contact("Bob", urns=["telegram:234567"], fields={"age": 40})
        devs = self.create_group("Devs", [bob])

        self.joe.name = "Jo\02e Blow"
        self.joe.save(update_fields=("name",))

        telegram = self.create_channel("TG", "Telegram", "765432")

        # messages can't be older than org
        self.org.created_on = datetime(2016, 1, 2, 10, tzinfo=pytz.UTC)
        self.org.save(update_fields=("created_on",))

        flow = self.create_flow("Color Flow")
        msg1 = self.create_incoming_msg(
            self.joe, "hello 1", created_on=datetime(2017, 1, 1, 10, tzinfo=pytz.UTC), flow=flow
        )
        msg2 = self.create_incoming_msg(
            bob, "hello 2", created_on=datetime(2017, 1, 2, 10, tzinfo=pytz.UTC), channel=telegram
        )
        msg3 = self.create_incoming_msg(
            bob, "hello 3", created_on=datetime(2017, 1, 3, 10, tzinfo=pytz.UTC), channel=telegram
        )

        # inbound message that looks like a surveyor message
        msg4 = self.create_incoming_msg(
            self.joe, "hello 4", surveyor=True, created_on=datetime(2017, 1, 4, 10, tzinfo=pytz.UTC)
        )

        # inbound message with media attached, such as an ivr recording
        msg5 = self.create_incoming_msg(
            self.joe,
            "Media message",
            attachments=["audio:http://rapidpro.io/audio/sound.mp3"],
            created_on=datetime(2017, 1, 5, 10, tzinfo=pytz.UTC),
        )

        # create some outbound messages with different statuses
        msg6 = self.create_outgoing_msg(
            self.joe, "Hey out 6", status=Msg.STATUS_SENT, created_on=datetime(2017, 1, 6, 10, tzinfo=pytz.UTC)
        )
        msg7 = self.create_outgoing_msg(
            bob,
            "Hey out 7",
            status=Msg.STATUS_DELIVERED,
            created_on=datetime(2017, 1, 7, 10, tzinfo=pytz.UTC),
            channel=telegram,
        )
        msg8 = self.create_outgoing_msg(
            self.joe, "Hey out 8", status=Msg.STATUS_ERRORED, created_on=datetime(2017, 1, 8, 10, tzinfo=pytz.UTC)
        )
        msg9 = self.create_outgoing_msg(
            self.joe, "Hey out 9", status=Msg.STATUS_FAILED, created_on=datetime(2017, 1, 9, 10, tzinfo=pytz.UTC)
        )

        self.assertEqual(msg5.get_attachments(), [Attachment("audio", "http://rapidpro.io/audio/sound.mp3")])

        # label first message
        label = self.create_label("la\02bel1")
        label.toggle_label([msg1], add=True)

        # archive last message
        msg3.visibility = Msg.VISIBILITY_ARCHIVED
        msg3.save()

        # create a dummy export task so that we won't be able to export
        blocking_export = ExportMessagesTask.create(
            self.org,
            self.admin,
            start_date=date(2017, 1, 1),
            end_date=date.today(),
            system_label=SystemLabel.TYPE_INBOX,
        )

        old_modified_on = blocking_export.modified_on

        response = self.client.post(
            export_url + "?l=I",
            {"export_all": 1, "start_date": "2022-09-01", "end_date": "2022-09-28"},
        )
        self.assertModalResponse(response, redirect="/msg/")

        response = self.client.get("/msg/")
        self.assertContains(response, "already an export in progress")

        # perform the export manually, assert how many queries
        with self.mockReadOnly():
            blocking_export.perform()

        blocking_export.refresh_from_db()
        # after performing the export `modified_on` should be updated
        self.assertNotEqual(old_modified_on, blocking_export.modified_on)

        def request_export(query, data=None):
            with self.mockReadOnly(assert_models={Msg, Contact}):
                response = self.client.post(export_url + query, data)
            self.assertModalResponse(response, redirect="/msg/")
            task = ExportMessagesTask.objects.order_by("-id").first()
            filename = "%s/test_orgs/%d/message_exports/%s.xlsx" % (settings.MEDIA_ROOT, self.org.id, task.uuid)
            workbook = load_workbook(filename=filename)
            return workbook.worksheets[0]

        expected_headers = [
            "Date",
            "Contact UUID",
            "Contact Name",
            "URN Scheme",
            "URN Value",
            "Flow",
            "Direction",
            "Text",
            "Attachments",
            "Status",
            "Channel",
            "Labels",
        ]

        # export all visible messages (i.e. not msg3) using export_all param
        with self.assertNumQueries(32):
            self.assertExcelSheet(
                request_export("?l=I", {"export_all": 1, "start_date": "2000-09-01", "end_date": "2022-09-28"}),
                [
                    expected_headers,
                    [
                        msg1.created_on,
                        msg1.contact.uuid,
                        "Joe Blow",
                        "tel",
                        "123",
                        "Color Flow",
                        "IN",
                        "hello 1",
                        "",
                        "handled",
                        "Test Channel",
                        "label1",
                    ],
                    [
                        msg2.created_on,
                        msg2.contact.uuid,
                        "Bob",
                        "telegram",
                        "234567",
                        "",
                        "IN",
                        "hello 2",
                        "",
                        "handled",
                        "Telegram",
                        "",
                    ],
                    [
                        msg4.created_on,
                        msg4.contact.uuid,
                        "Joe Blow",
                        "",
                        "",
                        "",
                        "IN",
                        "hello 4",
                        "",
                        "handled",
                        "",
                        "",
                    ],
                    [
                        msg5.created_on,
                        msg5.contact.uuid,
                        "Joe Blow",
                        "tel",
                        "123",
                        "",
                        "IN",
                        "Media message",
                        "http://rapidpro.io/audio/sound.mp3",
                        "handled",
                        "Test Channel",
                        "",
                    ],
                    [
                        msg6.created_on,
                        msg6.contact.uuid,
                        "Joe Blow",
                        "tel",
                        "123",
                        "",
                        "OUT",
                        "Hey out 6",
                        "",
                        "sent",
                        "Test Channel",
                        "",
                    ],
                    [
                        msg7.created_on,
                        msg7.contact.uuid,
                        "Bob",
                        "telegram",
                        "234567",
                        "",
                        "OUT",
                        "Hey out 7",
                        "",
                        "delivered",
                        "Telegram",
                        "",
                    ],
                    [
                        msg8.created_on,
                        msg8.contact.uuid,
                        "Joe Blow",
                        "tel",
                        "123",
                        "",
                        "OUT",
                        "Hey out 8",
                        "",
                        "errored",
                        "Test Channel",
                        "",
                    ],
                    [
                        msg9.created_on,
                        msg9.contact.uuid,
                        "Joe Blow",
                        "tel",
                        "123",
                        "",
                        "OUT",
                        "Hey out 9",
                        "",
                        "failed",
                        "Test Channel",
                        "",
                    ],
                ],
                self.org.timezone,
            )

        # check that notifications were created
        export = ExportMessagesTask.objects.order_by("id").last()
        self.assertEqual(
            1,
            self.admin.notifications.filter(
                notification_type="export:finished", message_export=export, email_status="P"
            ).count(),
        )

        # export just archived messages
        self.assertExcelSheet(
            request_export("?l=A", {"export_all": 0, "start_date": "2000-09-01", "end_date": "2022-09-28"}),
            [
                expected_headers,
                [
                    msg3.created_on,
                    msg3.contact.uuid,
                    "Bob",
                    "telegram",
                    "234567",
                    "",
                    "IN",
                    "hello 3",
                    "",
                    "handled",
                    "Telegram",
                    "",
                ],
            ],
            self.org.timezone,
        )

        # filter page should have an export option
        filter_url = reverse("msgs.msg_filter", args=[label.uuid])
        response = self.requestView(filter_url, self.admin)
        self.assertContains(response, label.name)
        self.assertContentMenu(filter_url, self.admin, ["Edit", "Download", "Usages", "Delete"])

        # try export with user label
        self.assertExcelSheet(
            request_export(
                "?l=%s" % label.uuid, {"export_all": 0, "start_date": "2000-09-01", "end_date": "2022-09-28"}
            ),
            [
                expected_headers,
                [
                    msg1.created_on,
                    msg1.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "Color Flow",
                    "IN",
                    "hello 1",
                    "",
                    "handled",
                    "Test Channel",
                    "label1",
                ],
            ],
            self.org.timezone,
        )

        # try export with a date range, a field and a group
        export_data = {
            "export_all": 1,
            "start_date": msg5.created_on.strftime("%Y-%m-%d"),
            "end_date": msg7.created_on.strftime("%Y-%m-%d"),
            "with_fields": [age.id],
            "with_groups": [devs.id],
        }

        self.assertExcelSheet(
            request_export("?l=I", export_data),
            [
                [
                    "Date",
                    "Contact UUID",
                    "Contact Name",
                    "URN Scheme",
                    "URN Value",
                    "Field:Age",
                    "Group:Devs",
                    "Flow",
                    "Direction",
                    "Text",
                    "Attachments",
                    "Status",
                    "Channel",
                    "Labels",
                ],
                [
                    msg5.created_on,
                    msg5.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    False,
                    "",
                    "IN",
                    "Media message",
                    "http://rapidpro.io/audio/sound.mp3",
                    "handled",
                    "Test Channel",
                    "",
                ],
                [
                    msg6.created_on,
                    msg6.contact.uuid,
                    "Joe Blow",
                    "tel",
                    "123",
                    "",
                    False,
                    "",
                    "OUT",
                    "Hey out 6",
                    "",
                    "sent",
                    "Test Channel",
                    "",
                ],
                [
                    msg7.created_on,
                    msg7.contact.uuid,
                    "Bob",
                    "telegram",
                    "234567",
                    "40",
                    True,
                    "",
                    "OUT",
                    "Hey out 7",
                    "",
                    "delivered",
                    "Telegram",
                    "",
                ],
            ],
            self.org.timezone,
        )

        # try to submit an invalid date (UI doesn't actually allow this)
        response = self.client.post(export_url + "?l=I", {"export_all": 1, "start_date": "xyz"})
        self.assertFormError(response, "form", "start_date", "Enter a valid date.")

        # try to submit without specifying dates (UI doesn't actually allow this)
        response = self.client.post(export_url + "?l=I", {"export_all": 1})
        self.assertFormError(response, "form", "start_date", "This field is required.")
        self.assertFormError(response, "form", "end_date", "This field is required.")

        # try to submit with start date in future
        response = self.client.post(
            export_url + "?l=I", {"export_all": 1, "start_date": "2200-01-01", "end_date": "2022-09-28"}
        )
        self.assertFormError(response, "form", None, "Start date can't be in the future.")

        # try to submit with start date > end date
        response = self.client.post(
            export_url + "?l=I", {"export_all": 1, "start_date": "2022-09-01", "end_date": "2022-03-01"}
        )
        self.assertFormError(response, "form", None, "End date can't be before start date.")

        # test as anon org to check that URNs don't end up in exports
        with AnonymousOrg(self.org):
            self.assertExcelSheet(
                request_export("?l=I", {"export_all": 1, "start_date": "2000-09-01", "end_date": "2022-09-28"}),
                [
                    [
                        "Date",
                        "Contact UUID",
                        "Contact Name",
                        "URN Scheme",
                        "Anon Value",
                        "Flow",
                        "Direction",
                        "Text",
                        "Attachments",
                        "Status",
                        "Channel",
                        "Labels",
                    ],
                    [
                        msg1.created_on,
                        msg1.contact.uuid,
                        "Joe Blow",
                        "tel",
                        self.joe.anon_display,
                        "Color Flow",
                        "IN",
                        "hello 1",
                        "",
                        "handled",
                        "Test Channel",
                        "label1",
                    ],
                    [
                        msg2.created_on,
                        msg2.contact.uuid,
                        "Bob",
                        "telegram",
                        bob.anon_display,
                        "",
                        "IN",
                        "hello 2",
                        "",
                        "handled",
                        "Telegram",
                        "",
                    ],
                    [
                        msg4.created_on,
                        msg4.contact.uuid,
                        "Joe Blow",
                        "",
                        self.joe.anon_display,
                        "",
                        "IN",
                        "hello 4",
                        "",
                        "handled",
                        "",
                        "",
                    ],
                    [
                        msg5.created_on,
                        msg5.contact.uuid,
                        "Joe Blow",
                        "tel",
                        self.joe.anon_display,
                        "",
                        "IN",
                        "Media message",
                        "http://rapidpro.io/audio/sound.mp3",
                        "handled",
                        "Test Channel",
                        "",
                    ],
                    [
                        msg6.created_on,
                        msg6.contact.uuid,
                        "Joe Blow",
                        "tel",
                        self.joe.anon_display,
                        "",
                        "OUT",
                        "Hey out 6",
                        "",
                        "sent",
                        "Test Channel",
                        "",
                    ],
                    [
                        msg7.created_on,
                        msg7.contact.uuid,
                        "Bob",
                        "telegram",
                        bob.anon_display,
                        "",
                        "OUT",
                        "Hey out 7",
                        "",
                        "delivered",
                        "Telegram",
                        "",
                    ],
                    [
                        msg8.created_on,
                        msg8.contact.uuid,
                        "Joe Blow",
                        "tel",
                        self.joe.anon_display,
                        "",
                        "OUT",
                        "Hey out 8",
                        "",
                        "errored",
                        "Test Channel",
                        "",
                    ],
                    [
                        msg9.created_on,
                        msg9.contact.uuid,
                        "Joe Blow",
                        "tel",
                        self.joe.anon_display,
                        "",
                        "OUT",
                        "Hey out 9",
                        "",
                        "failed",
                        "Test Channel",
                        "",
                    ],
                ],
                self.org.timezone,
            )

        response = self.client.post(
            export_url + "?l=I&redirect=http://foo.me",
            {"export_all": 1, "start_date": "2000-09-01", "end_date": "2022-09-28"},
        )
        self.assertModalResponse(response, redirect="/msg/")

        self.clear_storage()

    def test_fail_old_messages(self):
        msg1 = self.create_outgoing_msg(self.joe, "Hello", status=Msg.STATUS_QUEUED)
        msg2 = self.create_outgoing_msg(
            self.joe, "Hello", status=Msg.STATUS_QUEUED, created_on=timezone.now() - timedelta(days=8)
        )
        msg3 = self.create_outgoing_msg(
            self.joe, "Hello", status=Msg.STATUS_ERRORED, created_on=timezone.now() - timedelta(days=8)
        )
        msg4 = self.create_outgoing_msg(
            self.joe, "Hello", status=Msg.STATUS_SENT, created_on=timezone.now() - timedelta(days=8)
        )

        fail_old_messages()

        def assert_status(msg, status):
            msg.refresh_from_db()
            self.assertEqual(status, msg.status)

        assert_status(msg1, Msg.STATUS_QUEUED)
        assert_status(msg2, Msg.STATUS_FAILED)
        assert_status(msg3, Msg.STATUS_FAILED)
        assert_status(msg4, Msg.STATUS_SENT)

    def test_big_ids(self):
        # create an incoming message with big id
        msg = Msg.objects.create(
            id=3_000_000_000,
            org=self.org,
            direction="I",
            contact=self.joe,
            contact_urn=self.joe.urns.first(),
            text="Hi there",
            channel=self.channel,
            status="H",
            msg_type="T",
            visibility="V",
            created_on=timezone.now(),
        )
        ChannelLog.objects.create(
            id=3_000_000_000, channel=msg.channel, msg=msg, is_error=True, log_type=ChannelLog.LOG_TYPE_MSG_RECEIVE
        )
        spam = self.create_label("Spam")
        msg.labels.add(spam)

    def test_foreign_keys(self):
        # create a message which references a flow and a ticket
        flow = self.create_flow("Flow")
        contact = self.create_contact("Ann", phone="+250788000001")
        ticket = self.create_ticket(self.org.ticketers.get(), contact, "Help")
        msg = self.create_outgoing_msg(contact, "Hi", flow=flow, ticket=ticket)

        # both Msg.flow and Msg.ticket are unconstrained so we shuld be able to delete these
        flow.release(self.admin)
        flow.delete()
        ticket.delete()

        msg.refresh_from_db()

        # but then accessing them blows up
        with self.assertRaises(Flow.DoesNotExist):
            print(msg.flow)
        with self.assertRaises(Ticket.DoesNotExist):
            print(msg.ticket)


class MsgCRUDLTest(TembaTest, CRUDLTestMixin):
    def test_inbox(self):
        contact1 = self.create_contact("Joe Blow", phone="+250788000001")
        contact2 = self.create_contact("Frank", phone="+250788000002")
        msg1 = self.create_incoming_msg(contact1, "message number 1")
        msg2 = self.create_incoming_msg(contact1, "message number 2")
        msg3 = self.create_incoming_msg(contact2, "message number 3")
        msg4 = self.create_incoming_msg(contact2, "message number 4")
        msg5 = self.create_incoming_msg(contact2, "message number 5", visibility="A")
        self.create_incoming_msg(contact2, "message number 6", status=Msg.STATUS_PENDING)
        ChannelLog.objects.create(channel=self.channel, msg=msg1, log_type=ChannelLog.LOG_TYPE_MSG_RECEIVE)
        ChannelLog.objects.create(channel=self.channel, msg=msg2, log_type=ChannelLog.LOG_TYPE_MSG_RECEIVE)

        inbox_url = reverse("msgs.msg_inbox")

        # check query count
        self.login(self.admin)
        with self.assertNumQueries(14):
            self.client.get(inbox_url)

        response = self.assertListFetch(
            inbox_url + "?refresh=10000",
            allow_viewers=True,
            allow_editors=True,
            context_objects=[msg4, msg3, msg2, msg1],
        )

        # make sure that we embed refresh script if View.refresh is set
        self.assertContains(response, "function refresh")

        self.assertEqual(20000, response.context["refresh"])
        self.assertEqual(("archive", "label"), response.context["actions"])
        self.assertEqual({"count": 4, "label": "Inbox", "url": "/msg/"}, response.context["folders"][0])

        # test searching
        response = self.client.get(inbox_url + "?search=joe")
        self.assertEqual([msg2, msg1], list(response.context_data["object_list"]))

        # add some labels
        label1 = self.create_label("label1")
        self.create_label("label2")
        label3 = self.create_label("label3")

        # viewers can't label messages
        response = self.requestView(
            inbox_url, self.user, post_data={"action": "label", "objects": [msg1.id], "label": label1.id, "add": True}
        )
        self.assertEqual(403, response.status_code)

        # but editors can
        response = self.requestView(
            inbox_url,
            self.editor,
            post_data={"action": "label", "objects": [msg1.id, msg2.id], "label": label1.id, "add": True},
        )
        self.assertEqual(200, response.status_code)
        self.assertEqual({msg1, msg2}, set(label1.msgs.all()))

        # and remove labels
        self.requestView(
            inbox_url,
            self.editor,
            post_data={"action": "label", "objects": [msg2.id], "label": label1.id, "add": False},
        )
        self.assertEqual({msg1}, set(label1.msgs.all()))

        # can't label without a label object
        response = self.requestView(
            inbox_url,
            self.editor,
            post_data={"action": "label", "objects": [msg2.id], "add": False},
        )
        self.assertEqual({msg1}, set(label1.msgs.all()))

        # label more messages as admin
        self.requestView(
            inbox_url,
            self.admin,
            post_data={"action": "label", "objects": [msg1.id, msg2.id, msg3.id], "label": label3.id, "add": True},
        )
        self.assertEqual({msg1}, set(label1.msgs.all()))
        self.assertEqual({msg1, msg2, msg3}, set(label3.msgs.all()))

        # test archiving a msg
        self.client.post(inbox_url, {"action": "archive", "objects": msg1.id})
        self.assertEqual({msg1, msg5}, set(Msg.objects.filter(visibility=Msg.VISIBILITY_ARCHIVED)))

        # archiving doesn't remove labels
        msg1.refresh_from_db()
        self.assertEqual({label1, label3}, set(msg1.labels.all()))

        self.assertContentMenu(inbox_url, self.user, ["Download"])
        self.assertContentMenu(inbox_url, self.admin, ["Send Message", "New Label", "Download"])

    def test_flows(self):
        flow = self.create_flow("Test")
        contact1 = self.create_contact("Joe Blow", phone="+250788000001")
        msg1 = self.create_incoming_msg(contact1, "test 1", status="H", flow=flow)
        msg2 = self.create_incoming_msg(contact1, "test 2", status="H", flow=flow)
        self.create_incoming_msg(contact1, "test 3", status="H", flow=None)
        self.create_incoming_msg(contact1, "test 4", status="P", flow=None)

        flows_url = reverse("msgs.msg_flow")

        # check query count
        self.login(self.admin)
        with self.assertNumQueries(14):
            self.client.get(flows_url)

        response = self.assertListFetch(
            flows_url, allow_viewers=True, allow_editors=True, context_objects=[msg2, msg1]
        )

        self.assertEqual(("archive", "label"), response.context["actions"])

    def test_archived(self):
        contact1 = self.create_contact("Joe Blow", phone="+250788000001")
        contact2 = self.create_contact("Frank", phone="+250788000002")
        msg1 = self.create_incoming_msg(contact1, "message number 1", visibility=Msg.VISIBILITY_ARCHIVED)
        msg2 = self.create_incoming_msg(contact1, "message number 2", visibility=Msg.VISIBILITY_ARCHIVED)
        msg3 = self.create_incoming_msg(contact2, "message number 3", visibility=Msg.VISIBILITY_ARCHIVED)
        msg4 = self.create_incoming_msg(contact2, "message number 4", visibility=Msg.VISIBILITY_DELETED_BY_USER)
        self.create_incoming_msg(contact2, "message number 5", status=Msg.STATUS_PENDING)
        ChannelLog.objects.create(channel=self.channel, msg=msg1, log_type=ChannelLog.LOG_TYPE_MSG_RECEIVE)
        ChannelLog.objects.create(channel=self.channel, msg=msg2, log_type=ChannelLog.LOG_TYPE_MSG_RECEIVE)

        archived_url = reverse("msgs.msg_archived")

        # check query count
        self.login(self.admin)
        with self.assertNumQueries(14):
            self.client.get(archived_url)

        response = self.assertListFetch(
            archived_url + "?refresh=10000", allow_viewers=True, allow_editors=True, context_objects=[msg3, msg2, msg1]
        )

        self.assertEqual(("restore", "label", "delete"), response.context["actions"])
        self.assertEqual({"count": 3, "label": "Archived", "url": "/msg/archived/"}, response.context["folders"][2])

        # test searching
        response = self.client.get(archived_url + "?search=joe")
        self.assertEqual([msg2, msg1], list(response.context_data["object_list"]))

        # viewers can't restore messages
        response = self.requestView(archived_url, self.user, post_data={"action": "restore", "objects": [msg1.id]})
        self.assertEqual(403, response.status_code)

        # but editors can
        response = self.requestView(archived_url, self.editor, post_data={"action": "restore", "objects": [msg1.id]})
        self.assertEqual(200, response.status_code)
        self.assertEqual({msg2, msg3}, set(Msg.objects.filter(visibility=Msg.VISIBILITY_ARCHIVED)))

        # can also permanently delete messages
        response = self.requestView(archived_url, self.admin, post_data={"action": "delete", "objects": [msg2.id]})
        self.assertEqual(200, response.status_code)
        self.assertEqual({msg2, msg4}, set(Msg.objects.filter(visibility=Msg.VISIBILITY_DELETED_BY_USER)))
        self.assertEqual({msg3}, set(Msg.objects.filter(visibility=Msg.VISIBILITY_ARCHIVED)))

    def test_outbox(self):
        contact1 = self.create_contact("", phone="+250788382382")
        contact2 = self.create_contact("Joe Blow", phone="+250788000001")
        contact3 = self.create_contact("Frank Blow", phone="+250788000002")

        # create a single message broadcast that's sent but it's message is still not sent
        broadcast1 = self.create_broadcast(
            self.admin,
            "How is it going?",
            contacts=[contact1],
            status=Broadcast.STATUS_SENT,
            msg_status=Msg.STATUS_INITIALIZING,
        )
        msg1 = broadcast1.msgs.get()

        outbox_url = reverse("msgs.msg_outbox")

        # check query count
        self.login(self.admin)
        with self.assertNumQueries(14):
            self.client.get(outbox_url)

        # messages sorted by created_on
        response = self.assertListFetch(outbox_url, allow_viewers=True, allow_editors=True, context_objects=[msg1])

        self.assertEqual((), response.context["actions"])
        self.assertEqual({"count": 1, "label": "Outbox", "url": "/msg/outbox/"}, response.context["folders"][3])

        # create another broadcast this time with 3 messages
        contact4 = self.create_contact("Kevin", phone="+250788000003")
        group = self.create_group("Testers", contacts=[contact2, contact3])
        broadcast2 = self.create_broadcast(
            self.admin, "kLab is awesome", contacts=[contact4], groups=[group], msg_status=Msg.STATUS_QUEUED
        )
        msg4, msg3, msg2 = broadcast2.msgs.order_by("-id")

        broadcast3 = Broadcast.create(self.channel.org, self.admin, {"eng": "Pending broadcast"}, contacts=[contact4])
        broadcast4 = Broadcast.create(
            self.channel.org, self.admin, {"eng": "Scheduled broadcast"}, contacts=[contact4]
        )

        broadcast4.schedule = Schedule.create_schedule(self.org, self.admin, timezone.now(), Schedule.REPEAT_DAILY)
        broadcast4.save(update_fields=("schedule",))

        response = self.assertListFetch(
            outbox_url, allow_viewers=True, allow_editors=True, context_objects=[msg4, msg3, msg2, msg1]
        )

        # should see queued broadcast but not the scheduled one
        self.assertEqual(5, response.context_data["folders"][3]["count"])  # Outbox (includes queued broadcast)
        self.assertEqual([broadcast3], list(response.context_data["pending_broadcasts"]))

        response = self.client.get(outbox_url + "?search=kevin")
        self.assertEqual([Msg.objects.get(contact=contact4)], list(response.context_data["object_list"]))

        response = self.client.get(outbox_url + "?search=joe")
        self.assertEqual([Msg.objects.get(contact=contact2)], list(response.context_data["object_list"]))

        response = self.client.get(outbox_url + "?search=frank")
        self.assertEqual([Msg.objects.get(contact=contact3)], list(response.context_data["object_list"]))

        response = self.client.get(outbox_url + "?search=just")
        self.assertEqual([], list(response.context_data["object_list"]))

        response = self.client.get(outbox_url + "?search=klab")
        self.assertEqual([msg4, msg3, msg2], list(response.context_data["object_list"]))

    def test_sent(self):
        contact1 = self.create_contact("Joe Blow", phone="+250788000001")
        contact2 = self.create_contact("Frank Blow", phone="+250788000002")
        msg1 = self.create_outgoing_msg(contact1, "Hi 1", status="W", sent_on=timezone.now() - timedelta(hours=1))
        msg2 = self.create_outgoing_msg(contact1, "Hi 2", status="S", sent_on=timezone.now() - timedelta(hours=3))
        msg3 = self.create_outgoing_msg(contact2, "Hi 3", status="D", sent_on=timezone.now() - timedelta(hours=2))
        ChannelLog.objects.create(channel=self.channel, msg=msg1, log_type=ChannelLog.LOG_TYPE_MSG_SEND)
        ChannelLog.objects.create(channel=self.channel, msg=msg2, log_type=ChannelLog.LOG_TYPE_MSG_SEND)

        sent_url = reverse("msgs.msg_sent")

        # check query count
        self.login(self.admin)
        with self.assertNumQueries(13):
            self.client.get(sent_url)

        # messages sorted by sent_on
        response = self.assertListFetch(
            sent_url, allow_viewers=True, allow_editors=True, context_objects=[msg1, msg3, msg2]
        )

        self.assertContains(response, reverse("channels.channellog_msg", args=[msg1.channel.uuid, msg1.id]))

        response = self.client.get(sent_url + "?search=joe")
        self.assertEqual([msg1, msg2], list(response.context_data["object_list"]))

    @patch("temba.mailroom.client.MailroomClient.msg_resend")
    def test_failed(self, mock_msg_resend):
        contact1 = self.create_contact("Joe Blow", phone="+250788000001")
        msg1 = self.create_outgoing_msg(contact1, "message number 1", status="F")
        ChannelLog.objects.create(channel=msg1.channel, msg=msg1, is_error=True, log_type=ChannelLog.LOG_TYPE_MSG_SEND)

        failed_url = reverse("msgs.msg_failed")

        # create broadcast and fail the only message
        broadcast = self.create_broadcast(self.admin, "message number 2", contacts=[contact1])
        broadcast.get_messages().update(status="F")
        msg2 = broadcast.get_messages()[0]

        # message without a broadcast
        msg3 = self.create_outgoing_msg(contact1, "messsage number 3", status="F")

        # check query count
        self.login(self.admin)
        with self.assertNumQueries(13):
            self.client.get(failed_url)

        response = self.assertListFetch(
            failed_url, allow_viewers=True, allow_editors=True, context_objects=[msg3, msg2, msg1]
        )

        self.assertEqual(("resend",), response.context["actions"])
        self.assertContains(response, reverse("channels.channellog_msg", args=[msg1.channel.uuid, msg1.id]))

        # resend some messages
        self.client.post(failed_url, {"action": "resend", "objects": [msg2.id]})

        mock_msg_resend.assert_called_once_with(self.org.id, [msg2.id])

        # suspended orgs don't see resend as option
        self.org.is_suspended = True
        self.org.save(update_fields=("is_suspended",))

        response = self.client.get(failed_url)
        self.assertNotIn("resend", response.context["actions"])

    def test_filter(self):
        flow = self.create_flow("Flow")
        joe = self.create_contact("Joe Blow", phone="+250788000001")
        frank = self.create_contact("Frank Blow", phone="250788000002")
        billy = self.create_contact("Billy Bob", urns=["twitter:billy_bob"])

        # create labels
        label1 = self.create_label("label1")
        label2 = self.create_label("label2")
        label3 = self.create_label("label3")

        # create some messages
        msg1 = self.create_incoming_msg(joe, "test1")
        msg2 = self.create_incoming_msg(frank, "test2")
        msg3 = self.create_incoming_msg(billy, "test3")
        msg4 = self.create_incoming_msg(joe, "test4", visibility=Msg.VISIBILITY_ARCHIVED)
        msg5 = self.create_incoming_msg(joe, "test5", visibility=Msg.VISIBILITY_DELETED_BY_USER)
        msg6 = self.create_incoming_msg(joe, "IVR test", flow=flow)

        # apply the labels
        label1.toggle_label([msg1, msg2], add=True)
        label2.toggle_label([msg2, msg3], add=True)
        label3.toggle_label([msg1, msg2, msg3, msg4, msg5, msg6], add=True)

        label1_url = reverse("msgs.msg_filter", args=[label1.uuid])
        label3_url = reverse("msgs.msg_filter", args=[label3.uuid])

        # can't visit a filter page as a non-org user
        response = self.requestView(label3_url, self.non_org_user)
        self.assertRedirect(response, reverse("orgs.org_choose"))

        # can as org viewer user
        response = self.requestView(label3_url, self.user, HTTP_TEMBA_SPA=1)
        self.assertEqual(f"/msg/labels/{label3.uuid}", response.headers[TEMBA_MENU_SELECTION])
        self.assertEqual(200, response.status_code)
        self.assertEqual(("label",), response.context["actions"])
        self.assertContentMenu(label3_url, self.user, ["Download", "Usages"])  # no update or delete

        # check that non-visible messages are excluded, and messages and ordered newest to oldest
        self.assertEqual([msg6, msg3, msg2, msg1], list(response.context["object_list"]))

        # search on label by contact name
        response = self.client.get(f"{label3_url}?search=joe")
        self.assertEqual({msg1, msg6}, set(response.context_data["object_list"]))

        # check admin users see edit and delete options for labels
        self.assertContentMenu(label1_url, self.admin, ["Edit", "Download", "Usages", "Delete"])


class BroadcastTest(TembaTest):
    def setUp(self):
        super().setUp()

        self.joe = self.create_contact("Joe Blow", phone="123")
        self.frank = self.create_contact("Frank Blow", phone="321")

        self.just_joe = self.create_group("Just Joe", [self.joe])

        self.joe_and_frank = self.create_group("Joe and Frank", [self.joe, self.frank])

        self.kevin = self.create_contact(name="Kevin Durant", phone="987")
        self.lucy = self.create_contact(name="Lucy M", urns=["twitter:lucy"])

        # a Twitter channel
        self.twitter = self.create_channel("TT", "Twitter", "nyaruka")

    def test_delete(self):
        flow = self.create_flow("Test")
        label = self.create_label("Labeled")

        # create some incoming messages
        msg_in1 = self.create_incoming_msg(self.joe, "Hello")
        self.create_incoming_msg(self.frank, "Bonjour")

        # create a broadcast which is a response to an incoming message
        self.create_broadcast(self.user, "Noted", contacts=[self.joe])

        # create a broadcast which is to several contacts
        broadcast2 = self.create_broadcast(
            self.user, "Very old broadcast", groups=[self.joe_and_frank], contacts=[self.kevin, self.lucy]
        )

        # give joe some flow messages
        self.create_outgoing_msg(self.joe, "what's your fav color?")
        msg_in3 = self.create_incoming_msg(self.joe, "red!", flow=flow)
        self.create_outgoing_msg(self.joe, "red is cool")

        # mark all outgoing messages as sent except broadcast #2 to Joe
        Msg.objects.filter(direction="O").update(status="S")
        broadcast2.msgs.filter(contact=self.joe).update(status="F")

        # label one of our messages
        msg_in1.labels.add(label)
        self.assertEqual(LabelCount.get_totals([label])[label], 1)

        self.assertEqual(SystemLabel.get_counts(self.org)[SystemLabel.TYPE_INBOX], 2)
        self.assertEqual(SystemLabel.get_counts(self.org)[SystemLabel.TYPE_FLOWS], 1)
        self.assertEqual(SystemLabel.get_counts(self.org)[SystemLabel.TYPE_SENT], 6)
        self.assertEqual(SystemLabel.get_counts(self.org)[SystemLabel.TYPE_FAILED], 1)

        today = timezone.now().date()
        self.assertEqual(ChannelCount.get_day_count(self.channel, ChannelCount.INCOMING_MSG_TYPE, today), 3)
        self.assertEqual(ChannelCount.get_day_count(self.channel, ChannelCount.OUTGOING_MSG_TYPE, today), 6)
        self.assertEqual(ChannelCount.get_day_count(self.twitter, ChannelCount.INCOMING_MSG_TYPE, today), 0)
        self.assertEqual(ChannelCount.get_day_count(self.twitter, ChannelCount.OUTGOING_MSG_TYPE, today), 1)

        # delete all our messages save for our flow incoming message
        for m in Msg.objects.exclude(id=msg_in3.id):
            m.delete()

        # broadcasts should be unaffected
        self.assertEqual(2, Broadcast.objects.count())

        # check system label counts have been updated
        self.assertEqual(0, SystemLabel.get_counts(self.org)[SystemLabel.TYPE_INBOX])
        self.assertEqual(1, SystemLabel.get_counts(self.org)[SystemLabel.TYPE_FLOWS])
        self.assertEqual(0, SystemLabel.get_counts(self.org)[SystemLabel.TYPE_SENT])
        self.assertEqual(0, SystemLabel.get_counts(self.org)[SystemLabel.TYPE_FAILED])

        # check user label
        self.assertEqual(0, LabelCount.get_totals([label])[label])

        # but daily channel counts should be unchanged
        self.assertEqual(3, ChannelCount.get_day_count(self.channel, ChannelCount.INCOMING_MSG_TYPE, today))
        self.assertEqual(6, ChannelCount.get_day_count(self.channel, ChannelCount.OUTGOING_MSG_TYPE, today))
        self.assertEqual(0, ChannelCount.get_day_count(self.twitter, ChannelCount.INCOMING_MSG_TYPE, today))
        self.assertEqual(1, ChannelCount.get_day_count(self.twitter, ChannelCount.OUTGOING_MSG_TYPE, today))

    def test_model(self):
        schedule = Schedule.create_schedule(self.org, self.admin, timezone.now(), Schedule.REPEAT_MONTHLY)
        broadcast1 = Broadcast.create(
            self.org,
            self.user,
            {"eng": "Hello everyone", "spa": "Hola a todos", "fra": "Salut à tous"},
            base_language="eng",
            groups=[self.joe_and_frank],
            contacts=[self.kevin, self.lucy],
            schedule=schedule,
        )
        self.assertEqual("Q", broadcast1.status)
        self.assertEqual(True, broadcast1.is_active)

        with patch("temba.mailroom.queue_broadcast") as mock_queue_broadcast:
            broadcast1.send_async()

            mock_queue_broadcast.assert_called_once_with(broadcast1)

        # create a broadcast that looks like it has been sent
        broadcast2 = self.create_broadcast(self.admin, "Hi everyone", contacts=[self.kevin, self.lucy])

        self.assertEqual(2, broadcast2.msgs.count())
        self.assertEqual(2, broadcast2.get_message_count())

        self.assertEqual(2, Broadcast.objects.count())
        self.assertEqual(2, Msg.objects.count())
        self.assertEqual(1, Schedule.objects.count())

        broadcast1.delete(self.admin, soft=True)

        self.assertEqual(2, Broadcast.objects.count())
        self.assertEqual(2, Msg.objects.count())
        self.assertEqual(1, Schedule.objects.count())

        # schedule should also be inactive
        schedule.refresh_from_db()
        self.assertFalse(schedule.is_active)

        broadcast1.delete(self.admin, soft=False)
        broadcast2.delete(self.admin, soft=False)

        self.assertEqual(0, Broadcast.objects.count())
        self.assertEqual(0, Msg.objects.count())
        self.assertEqual(0, Schedule.objects.count())

        with self.assertRaises(AssertionError):
            Broadcast.create(self.org, self.user, "no recipients")

    def test_get_translation(self):
        # create a broadcast with 3 different languages containing both text and attachments
        eng_text = "Hello everyone"
        spa_text = "Hola a todos"
        fra_text = "Salut à tous"

        # create 3 attachments
        media_attachments = []
        for _ in range(3):
            media = Media.from_upload(
                self.org,
                self.admin,
                self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
                process=False,
            )
            media_attachments.append({"content_type": media.content_type, "url": media.url})
        attachments = compose_deserialize_attachments(media_attachments)
        eng_attachments = [attachments[0]]
        spa_attachments = [attachments[1]]
        fra_attachments = [attachments[2]]

        broadcast = Broadcast.create(
            self.org,
            self.user,
            text={"eng": eng_text, "spa": spa_text, "fra": fra_text},
            attachments={"eng": eng_attachments, "spa": spa_attachments, "fra": fra_attachments},
            base_language="eng",
            groups=[self.joe_and_frank],
            contacts=[self.kevin, self.lucy],
            schedule=Schedule.create_schedule(self.org, self.admin, timezone.now(), Schedule.REPEAT_MONTHLY),
        )

        self.org.set_flow_languages(self.admin, ["kin"])

        # uses broadcast base language
        self.assertEqual(eng_text, broadcast.get_translation(self.joe)["text"])
        self.assertEqual(eng_attachments, broadcast.get_translation(self.joe)["attachments"])

        self.org.set_flow_languages(self.admin, ["spa", "eng", "fra"])

        # uses org primary language
        self.assertEqual(spa_text, broadcast.get_translation(self.joe)["text"])
        self.assertEqual(spa_attachments, broadcast.get_translation(self.joe)["attachments"])

        self.joe.language = "fra"
        self.joe.save(update_fields=("language",))

        # uses contact language
        self.assertEqual(fra_text, broadcast.get_translation(self.joe)["text"])
        self.assertEqual(fra_attachments, broadcast.get_translation(self.joe)["attachments"])

        self.org.set_flow_languages(self.admin, ["spa", "eng"])

        # but only if it's allowed
        self.assertEqual(spa_text, broadcast.get_translation(self.joe)["text"])
        self.assertEqual(spa_attachments, broadcast.get_translation(self.joe)["attachments"])

        self.assertEqual(f'<Broadcast: id={broadcast.id} text="Hola a todos">', repr(broadcast))


class BroadcastCRUDLTest(TembaTest, CRUDLTestMixin):
    def setUp(self):
        super().setUp()

        self.joe = self.create_contact("Joe Blow", urns=["tel:+12025550149"])
        self.frank = self.create_contact("Frank Blow", urns=["tel:+12025550195"])
        self.joe_and_frank = self.create_group("Joe and Frank", [self.joe, self.frank])

    @patch("temba.mailroom.queue_broadcast")
    def test_send(self, mock_queue_broadcast):
        send_url = reverse("msgs.broadcast_send")

        # can't send if you're not logged in
        response = self.client.get(send_url)
        self.assertLoginRedirect(response)

        response = self.client.post(
            send_url, {"text": "Test", "omnibox": omnibox_serialize(self.org, [], [self.joe], json_encode=True)}
        )
        self.assertLoginRedirect(response)

        # or just a viewer user
        self.login(self.user)

        response = self.client.get(send_url)
        self.assertLoginRedirect(response)

        # but editors can
        self.login(self.editor)

        response = self.client.get(send_url)
        self.assertEqual(["omnibox", "text", "step_node", "loc"], list(response.context["form"].fields.keys()))

        # initialize form based on a contact
        response = self.client.get(f"{send_url}?c={self.joe.uuid}")
        omnibox = response.context["form"]["omnibox"]
        self.assertEqual(
            [{"id": str(self.joe.uuid), "name": "Joe Blow", "type": "contact", "urn": "(202) 555-0149"}],
            omnibox.value(),
        )

        # submit with a send to a group and a contact
        response = self.client.post(
            send_url,
            {
                "text": "Hey Joe, where you goin?",
                "omnibox": omnibox_serialize(self.org, [self.joe_and_frank], [self.frank], json_encode=True),
            },
        )
        self.assertEqual(200, response.status_code)

        broadcast = Broadcast.objects.get()
        self.assertEqual({"und": {"text": "Hey Joe, where you goin?"}}, broadcast.translations)
        self.assertEqual({self.joe_and_frank}, set(broadcast.groups.all()))
        self.assertEqual({self.frank}, set(broadcast.contacts.all()))

        mock_queue_broadcast.assert_called_once_with(broadcast)

        # try to submit a send to nobody
        response = self.client.post(
            send_url, {"text": "Broken", "omnibox": omnibox_serialize(self.org, [], [], json_encode=True)}
        )
        self.assertFormError(response, "form", "omnibox", "At least one recipient is required.")

        # if we release our send channel we also can't start send
        self.channel.release(self.admin)

        response = self.requestView(send_url, self.admin)
        self.assertContains(
            response, 'To get started you need to <a href="/channels/channel/claim/">add a channel</a>'
        )
        self.assertNotContains(response, "Send")

    @patch("temba.mailroom.queue_broadcast")
    def test_send_to_node(self, mock_queue_broadcast):
        send_url = reverse("msgs.broadcast_send")

        self.login(self.editor)

        # give Joe a flow run that has stopped on a node
        flow = self.get_flow("color_v13")
        flow_nodes = flow.get_definition()["nodes"]
        color_prompt = flow_nodes[0]
        color_split = flow_nodes[4]
        (
            MockSessionWriter(self.joe, flow)
            .visit(color_prompt)
            .send_msg("What is your favorite color?", self.channel)
            .visit(color_split)
            .wait()
            .save()
        ).session.runs.get()

        # initialize form based on a flow node UUID
        response = self.client.get(f"{send_url}?step_node={color_split['uuid']}")

        # no omnibox...
        self.assertEqual(["text", "step_node", "loc"], list(response.context["form"].fields.keys()))

        response = self.client.post(send_url, {"text": "Hurry up", "step_node": color_split["uuid"]})
        self.assertEqual("hide", response["Temba-Success"])

        broadcast = Broadcast.objects.get()
        self.assertEqual(broadcast.translations, {"und": {"text": "Hurry up"}})
        self.assertEqual(broadcast.groups.count(), 0)
        self.assertEqual({self.joe}, set(broadcast.contacts.all()))

        mock_queue_broadcast.assert_called_once_with(broadcast)

        # if there are no contacts at the given node, we don't actually create a broadcast
        response = self.client.post(
            send_url, {"text": "Hurry up", "step_node": "36b2c697-a1d9-47a9-9553-d07d6a725877"}
        )
        self.assertEqual("hide", response["Temba-Success"])

        self.assertEqual(1, Broadcast.objects.count())

    def test_scheduled(self):
        list_url = reverse("msgs.broadcast_scheduled")

        self.assertListFetch(list_url, allow_viewers=True, allow_editors=True, context_objects=[])
        self.assertContentMenu(list_url, self.user, [])
        self.assertContentMenu(list_url, self.admin, ["New Broadcast"])

        bc1 = self.create_broadcast(
            self.admin,
            "good morning",
            contacts=[self.joe],
            schedule=Schedule.create_schedule(self.org, self.admin, timezone.now(), Schedule.REPEAT_DAILY),
        )
        bc2 = self.create_broadcast(
            self.admin,
            "good evening",
            contacts=[self.frank],
            schedule=Schedule.create_schedule(self.org, self.admin, timezone.now(), Schedule.REPEAT_DAILY),
        )
        self.create_broadcast(self.admin, "not scheduled", groups=[self.joe_and_frank])

        bc3 = self.create_broadcast(
            self.admin,
            "good afternoon",
            contacts=[self.frank],
            schedule=Schedule.create_schedule(self.org, self.admin, timezone.now(), Schedule.REPEAT_DAILY),
        )

        self.assertListFetch(list_url, allow_viewers=True, allow_editors=True, context_objects=[bc3, bc2, bc1])

        bc3.is_active = False
        bc3.save(update_fields=("is_active",))

        self.assertListFetch(list_url, allow_viewers=True, allow_editors=True, context_objects=[bc2, bc1])

        # can search on text or URN path
        self.assertListFetch(list_url + "?search=MORN", allow_viewers=True, allow_editors=True, context_objects=[bc1])
        self.assertListFetch(list_url + "?search=50195", allow_viewers=True, allow_editors=True, context_objects=[bc2])

    def test_scheduled_create(self):
        create_url = reverse("msgs.broadcast_scheduled_create")

        self.assertCreateFetch(
            create_url,
            allow_viewers=False,
            allow_editors=True,
            form_fields=["start_datetime", "repeat_period", "repeat_days_of_week", "omnibox", "compose"],
        )

        # try to create with no values
        self.assertCreateSubmit(
            create_url,
            {
                "compose": compose_serialize(json_encode=True),
            },
            form_errors={
                "omnibox": "At least one recipient is required.",
                "compose": "Text or attachments are required.",
                "start_datetime": "This field is required.",
                "repeat_period": "This field is required.",
            },
        )

        # create with only text
        text = "Daily reminder"
        attachments = []
        translation = {"text": text, "attachments": attachments}
        response = self.assertCreateSubmit(
            create_url,
            {
                "omnibox": omnibox_serialize(self.org, groups=[], contacts=[self.joe, self.frank], json_encode=True),
                "compose": compose_serialize(translation, json_encode=True),
                "start_datetime": "2021-06-24 12:00",
                "repeat_period": "W",
                "repeat_days_of_week": ["M", "F"],
            },
            new_obj_query=Broadcast.objects.filter(
                translations={"und": {"text": "Daily reminder", "attachments": attachments}},
                schedule__repeat_period="W",
                schedule__repeat_days_of_week="MF",
            ),
            success_status=200,
        )
        self.assertEqual("/broadcast/scheduled/", response["Temba-Success"])

        # create an attachment
        media_attachments = []
        media1 = Media.from_upload(
            self.org,
            self.admin,
            self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
            process=False,
        )
        media_attachments.append({"content_type": media1.content_type, "url": media1.url})

        # create with only 1 attachment
        text = ""
        attachments = compose_deserialize_attachments(media_attachments)
        translation = {"text": text, "attachments": attachments}
        response = self.assertCreateSubmit(
            create_url,
            {
                "omnibox": omnibox_serialize(self.org, groups=[], contacts=[self.joe, self.frank], json_encode=True),
                "compose": compose_serialize(translation, json_encode=True),
                "start_datetime": "2021-06-24 12:00",
                "repeat_period": "W",
                "repeat_days_of_week": ["M", "F"],
            },
            new_obj_query=Broadcast.objects.filter(
                translations={"und": {"text": text, "attachments": attachments}},
                schedule__repeat_period="W",
                schedule__repeat_days_of_week="MF",
            ),
            success_status=200,
        )
        self.assertEqual("/broadcast/scheduled/", response["Temba-Success"])

        # create another attachment
        media2 = Media.from_upload(
            self.org,
            self.admin,
            self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
            process=False,
        )
        media_attachments.append({"content_type": media2.content_type, "url": media2.url})

        # create with text and 2 attachments
        text = "Daily reminder"
        attachments = compose_deserialize_attachments(media_attachments)
        translation = {"text": text, "attachments": attachments}
        response = self.assertCreateSubmit(
            create_url,
            {
                "omnibox": omnibox_serialize(self.org, groups=[], contacts=[self.joe, self.frank], json_encode=True),
                "compose": compose_serialize(translation, json_encode=True),
                "start_datetime": "2021-06-24 12:00",
                "repeat_period": "W",
                "repeat_days_of_week": ["M", "F"],
            },
            new_obj_query=Broadcast.objects.filter(
                translations={"und": {"text": text, "attachments": attachments}},
                schedule__repeat_period="W",
                schedule__repeat_days_of_week="MF",
            ),
            success_status=200,
        )
        self.assertEqual("/broadcast/scheduled/", response["Temba-Success"])

    def test_scheduled_create_text_max_length(self):
        create_url = reverse("msgs.broadcast_scheduled_create")
        omnibox = omnibox_serialize(self.org, groups=[], contacts=[self.joe, self.frank], json_encode=True)

        text = "".join(
            random.choices(string.ascii_letters + string.digits + string.punctuation, k=Msg.MAX_TEXT_LEN + 1)
        )
        attachments = []
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)

        self.assertCreateSubmit(
            create_url,
            {
                "omnibox": omnibox,
                "compose": compose,
                "start_datetime": "2021-06-24 12:00",
                "repeat_period": "W",
                "repeat_days_of_week": ["M", "F"],
            },
            form_errors={"compose": f"Maximum allowed text is {Msg.MAX_TEXT_LEN} characters."},
        )

    def test_scheduled_create_attachments_max_files(self):
        create_url = reverse("msgs.broadcast_scheduled_create")
        omnibox = omnibox_serialize(self.org, groups=[], contacts=[self.joe, self.frank], json_encode=True)

        media_attachments = []
        for _ in range(Msg.MAX_ATTACHMENTS + 1):
            media = Media.from_upload(
                self.org,
                self.admin,
                self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
                process=False,
            )
            media_attachments.append({"content_type": media.content_type, "url": media.url})
        text = ""
        attachments = compose_deserialize_attachments(media_attachments)
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)

        self.assertCreateSubmit(
            create_url,
            {
                "omnibox": omnibox,
                "compose": compose,
                "start_datetime": "2021-06-24 12:00",
                "repeat_period": "W",
                "repeat_days_of_week": ["M", "F"],
            },
            form_errors={"compose": f"Maximum allowed attachments is {Msg.MAX_ATTACHMENTS} files."},
        )

    def test_scheduled_read(self):
        schedule = Schedule.create_schedule(self.org, self.admin, timezone.now(), "D", repeat_days_of_week="MWF")
        broadcast = self.create_broadcast(
            self.admin,
            "Daily reminder",
            groups=[self.joe_and_frank],
            schedule=schedule,
        )

        read_url = reverse("msgs.broadcast_scheduled_read", args=[broadcast.id])

        self.login(self.editor)

        # view with empty send history
        response = self.client.get(read_url)
        self.assertEqual(broadcast, response.context["object"])
        self.assertEqual([], list(response.context["send_history"]))

        # add some send history
        sends = []
        for i in range(3):
            text = "Daily Reminder"
            media = Media.from_upload(
                self.org,
                self.admin,
                self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
                process=False,
            )
            attachment = {"content_type": media.content_type, "url": media.url}
            attachments = compose_deserialize_attachments([attachment])

            sends.append(
                Broadcast.create(
                    self.org,
                    self.admin,
                    text={"eng": text},
                    attachments={"eng": attachments},
                    groups=[self.joe_and_frank],
                    status=Msg.STATUS_QUEUED,
                    parent=broadcast,
                )
            )

        # sends are listed newest first
        response = self.client.get(read_url)
        self.assertEqual(response.context["object"], broadcast)
        self.assertEqual(list(reversed(sends)), list(response.context["send_history"]))

    def test_scheduled_update(self):
        # create a broadcast via send
        self.login(self.editor)
        omnibox = omnibox_serialize(self.org, [], [self.joe], json_encode=True)
        text = "Lunch reminder"
        attachments = []
        self.client.post(reverse("msgs.broadcast_send"), dict(omnibox=omnibox, text=text, schedule=True))
        broadcast = Broadcast.objects.get()
        url = reverse("msgs.broadcast_scheduled_update", args=[broadcast.pk])
        response = self.client.get(url)
        self.assertEqual(list(response.context["form"].fields.keys()), ["omnibox", "compose", "loc"])

        # update the broadcast's contact
        omnibox = omnibox_serialize(self.org, [], [self.frank], json_encode=True)
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)
        response = self.client.post(url, dict(omnibox=omnibox, compose=compose))
        self.assertEqual(response.status_code, 302)
        broadcast = Broadcast.objects.get()
        self.assertEqual(broadcast.translations, {"und": {"text": text, "attachments": attachments}})
        self.assertEqual(broadcast.base_language, "und")
        self.assertEqual(set(broadcast.contacts.all()), {self.frank})

        # update the broadcast's text
        text = "Dinner reminder"
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)
        response = self.client.post(url, dict(omnibox=omnibox, compose=compose))
        self.assertEqual(response.status_code, 302)
        broadcast = Broadcast.objects.get()
        self.assertEqual(broadcast.translations, {"und": {"text": text, "attachments": attachments}})
        self.assertEqual(broadcast.base_language, "und")
        self.assertEqual(set(broadcast.contacts.all()), {self.frank})

        # create an attachment and update the broadcast's attachments
        media_attachments = []
        media1 = Media.from_upload(
            self.org,
            self.admin,
            self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
            process=False,
        )
        media_attachments.append({"content_type": media1.content_type, "url": media1.url})

        attachments = compose_deserialize_attachments(media_attachments)
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)
        response = self.client.post(url, dict(omnibox=omnibox, compose=compose))
        self.assertEqual(response.status_code, 302)
        broadcast = Broadcast.objects.get()
        self.assertEqual(broadcast.translations, {"und": {"text": text, "attachments": attachments}})
        self.assertEqual(broadcast.base_language, "und")
        self.assertEqual(set(broadcast.contacts.all()), {self.frank})

        # create another attachment and update the broadcast's text and attachments
        media2 = Media.from_upload(
            self.org,
            self.admin,
            self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
            process=False,
        )
        media_attachments.append({"content_type": media2.content_type, "url": media2.url})

        text = "Midnight snack reminder"
        attachments = compose_deserialize_attachments(media_attachments)
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)
        response = self.client.post(url, dict(omnibox=omnibox, compose=compose))
        self.assertEqual(response.status_code, 302)
        broadcast = Broadcast.objects.get()
        self.assertEqual(broadcast.translations, {"und": {"text": text, "attachments": attachments}})
        self.assertEqual(broadcast.base_language, "und")
        self.assertEqual(set(broadcast.contacts.all()), {self.frank})

    def test_scheduled_update_missing_contacts(self):
        self.login(self.editor)

        # create a broadcast via send
        omnibox = omnibox_serialize(self.org, [self.joe_and_frank], [self.joe], json_encode=True)
        self.client.post(reverse("msgs.broadcast_send"), dict(omnibox=omnibox, text="Lunch reminder", schedule=True))
        broadcast = Broadcast.objects.get()

        # update the broadcast with no groups or contacts
        omnibox = omnibox_serialize(self.org, [], [], json_encode=True)
        text = "Empty contacts"
        attachments = []
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)
        response = self.client.post(
            reverse("msgs.broadcast_scheduled_update", args=[broadcast.pk]),
            dict(omnibox=omnibox, compose=compose, schedule=True),
        )
        self.assertFormError(response, "form", None, "At least one recipient is required.")

    def test_scheduled_update_missing_media(self):
        self.login(self.editor)

        # create a broadcast via send
        omnibox = omnibox_serialize(self.org, [self.joe_and_frank], [self.joe], json_encode=True)
        self.client.post(reverse("msgs.broadcast_send"), dict(omnibox=omnibox, text="Lunch reminder", schedule=True))
        broadcast = Broadcast.objects.get()

        # update the broadcast with no text or attachments
        compose = compose_serialize(json_encode=True)
        response = self.client.post(
            reverse("msgs.broadcast_scheduled_update", args=[broadcast.pk]),
            dict(omnibox=omnibox, compose=compose, schedule=True),
        )
        self.assertFormError(response, "form", None, "Text or attachments are required.")

    def test_scheduled_update_text_max_length(self):
        self.login(self.editor)

        # create a broadcast via send
        omnibox = omnibox_serialize(self.org, [self.joe_and_frank], [self.joe], json_encode=True)
        self.client.post(reverse("msgs.broadcast_send"), dict(omnibox=omnibox, text="Lunch reminder", schedule=True))
        broadcast = Broadcast.objects.get()

        # update the broadcast's text
        text = "".join(
            random.choices(string.ascii_letters + string.digits + string.punctuation, k=Msg.MAX_TEXT_LEN + 1)
        )
        attachments = []
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)

        response = self.client.post(
            reverse("msgs.broadcast_scheduled_update", args=[broadcast.pk]),
            dict(omnibox=omnibox, compose=compose, schedule=True),
        )
        self.assertFormError(response, "form", None, f"Maximum allowed text is {Msg.MAX_TEXT_LEN} characters.")

    def test_scheduled_update_attachments_max_files(self):
        self.login(self.editor)

        # create a broadcast via send
        omnibox = omnibox_serialize(self.org, [self.joe_and_frank], [self.joe], json_encode=True)
        self.client.post(reverse("msgs.broadcast_send"), dict(omnibox=omnibox, text="Lunch reminder", schedule=True))
        broadcast = Broadcast.objects.get()

        # update the broadcast's attachments
        media_attachments = []
        for _ in range(Msg.MAX_ATTACHMENTS + 1):
            media = Media.from_upload(
                self.org,
                self.admin,
                self.upload(f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg", "image/jpeg"),
                process=False,
            )
            media_attachments.append({"content_type": media.content_type, "url": media.url})

        text = ""
        attachments = compose_deserialize_attachments(media_attachments)
        translation = {"text": text, "attachments": attachments}
        compose = compose_serialize(translation, json_encode=True)

        response = self.client.post(
            reverse("msgs.broadcast_scheduled_update", args=[broadcast.pk]),
            dict(omnibox=omnibox, compose=compose, schedule=True),
        )
        self.assertFormError(response, "form", None, f"Maximum allowed attachments is {Msg.MAX_ATTACHMENTS} files.")

    def test_scheduled_delete(self):
        self.login(self.editor)
        schedule = Schedule.create_schedule(self.org, self.admin, timezone.now(), "D", repeat_days_of_week="MWF")
        broadcast = self.create_broadcast(
            self.admin,
            "Daily reminder",
            groups=[self.joe_and_frank],
            schedule=schedule,
        )

        delete_url = reverse("msgs.broadcast_scheduled_delete", args=[broadcast.id])

        # fetch the delete modal
        response = self.assertDeleteFetch(delete_url, allow_editors=True, as_modal=True)
        self.assertContains(response, "You are about to delete")

        # submit the delete modal
        response = self.assertDeleteSubmit(delete_url, object_deactivated=broadcast, success_status=200)
        self.assertEqual("/broadcast/scheduled/", response["Temba-Success"])

        broadcast = Broadcast.objects.get(id=broadcast.id)
        schedule = Schedule.objects.get(id=schedule.id)

        self.assertFalse(broadcast.is_active)


class LabelTest(TembaTest):
    def setUp(self):
        super().setUp()

        self.joe = self.create_contact("Joe Blow", phone="073835001")
        self.frank = self.create_contact("Frank", phone="073835002")

    def test_create(self):
        label1 = Label.create(self.org, self.user, "Spam")
        self.assertEqual("Spam", label1.name)
        self.assertIsNone(label1.folder)

        # don't allow invalid name
        self.assertRaises(AssertionError, Label.create, self.org, self.user, '"Hi"')

        # don't allow duplicate name
        self.assertRaises(AssertionError, Label.create, self.org, self.user, "Spam")

    def test_toggle_label(self):
        label = self.create_label("Spam")
        msg1 = self.create_incoming_msg(self.joe, "Message 1")
        msg2 = self.create_incoming_msg(self.joe, "Message 2")
        msg3 = self.create_incoming_msg(self.joe, "Message 3")

        self.assertEqual(label.get_visible_count(), 0)

        label.toggle_label([msg1, msg2, msg3], add=True)  # add label to 3 messages

        label.refresh_from_db()
        self.assertEqual(label.get_visible_count(), 3)
        self.assertEqual(set(label.get_messages()), {msg1, msg2, msg3})

        label.toggle_label([msg3], add=False)  # remove label from a message

        label.refresh_from_db()
        self.assertEqual(label.get_visible_count(), 2)
        self.assertEqual(set(label.get_messages()), {msg1, msg2})

        # check still correct after squashing
        squash_msg_counts()
        self.assertEqual(label.get_visible_count(), 2)

        msg2.archive()  # won't remove label from msg, but msg no longer counts toward visible count

        label.refresh_from_db()
        self.assertEqual(label.get_visible_count(), 1)
        self.assertEqual(set(label.get_messages()), {msg1, msg2})

        msg2.restore()  # msg back in visible count

        label.refresh_from_db()
        self.assertEqual(label.get_visible_count(), 2)
        self.assertEqual(set(label.get_messages()), {msg1, msg2})

        msg2.delete()  # removes label message no longer visible

        label.refresh_from_db()
        self.assertEqual(label.get_visible_count(), 1)
        self.assertEqual(set(label.get_messages()), {msg1})

        msg3.archive()
        label.toggle_label([msg3], add=True)  # labelling an already archived message doesn't increment the count

        label.refresh_from_db()
        self.assertEqual(label.get_visible_count(), 1)
        self.assertEqual(set(label.get_messages()), {msg1, msg3})

        msg3.restore()  # but then restoring that message will

        label.refresh_from_db()
        self.assertEqual(label.get_visible_count(), 2)
        self.assertEqual(set(label.get_messages()), {msg1, msg3})

        # can't label outgoing messages
        msg5 = self.create_outgoing_msg(self.joe, "Message")
        self.assertRaises(AssertionError, label.toggle_label, [msg5], add=True)

        # squashing shouldn't affect counts
        self.assertEqual(LabelCount.get_totals([label])[label], 2)

        squash_msg_counts()

        self.assertEqual(LabelCount.get_totals([label])[label], 2)

    def test_delete(self):
        label1 = self.create_label("Spam")
        label2 = self.create_label("Social")
        label3 = self.create_label("Other")

        msg1 = self.create_incoming_msg(self.joe, "Message 1")
        msg2 = self.create_incoming_msg(self.joe, "Message 2")
        msg3 = self.create_incoming_msg(self.joe, "Message 3")

        label1.toggle_label([msg1, msg2], add=True)
        label2.toggle_label([msg1], add=True)
        label3.toggle_label([msg3], add=True)

        ExportMessagesTask.create(self.org, self.admin, start_date=date.today(), end_date=date.today(), label=label1)

        label1.release(self.admin)
        label2.release(self.admin)

        # check that contained labels are also released
        self.assertEqual(0, Label.objects.filter(id__in=[label1.id, label2.id], is_active=True).count())
        self.assertEqual(set(), set(Msg.objects.get(id=msg1.id).labels.all()))
        self.assertEqual(set(), set(Msg.objects.get(id=msg2.id).labels.all()))
        self.assertEqual({label3}, set(Msg.objects.get(id=msg3.id).labels.all()))

        label3.release(self.admin)
        label3.refresh_from_db()

        self.assertFalse(label3.is_active)
        self.assertEqual(self.admin, label3.modified_by)
        self.assertEqual(set(), set(Msg.objects.get(id=msg3.id).labels.all()))


class LabelCRUDLTest(TembaTest, CRUDLTestMixin):
    def test_create(self):
        create_url = reverse("msgs.label_create")

        self.assertCreateFetch(create_url, allow_viewers=False, allow_editors=True, form_fields=("name", "messages"))

        # try to create label with invalid name
        self.assertCreateSubmit(
            create_url, {"name": '"Spam"'}, form_errors={"name": 'Cannot contain the character: "'}
        )

        # try again with valid name
        self.assertCreateSubmit(
            create_url,
            {"name": "Spam"},
            new_obj_query=Label.objects.filter(name="Spam"),
        )

        # check that we can't create another with same name
        self.assertCreateSubmit(create_url, {"name": "Spam"}, form_errors={"name": "Must be unique."})

        # create another label
        self.assertCreateSubmit(
            create_url,
            {"name": "Spam 2"},
            new_obj_query=Label.objects.filter(name="Spam 2"),
        )

        # try creating a new label after reaching the limit on labels
        current_count = Label.get_active_for_org(self.org).count()
        with override_settings(ORG_LIMIT_DEFAULTS={"labels": current_count}):
            response = self.client.post(create_url, {"name": "CoolStuff"})
            self.assertFormError(
                response,
                "form",
                "name",
                "This workspace has reached its limit of 2 labels. "
                "You must delete existing ones before you can create new ones.",
            )

    def test_update(self):
        label1 = self.create_label("Spam")
        label2 = self.create_label("Sales")

        label1_url = reverse("msgs.label_update", args=[label1.id])
        label2_url = reverse("msgs.label_update", args=[label2.id])

        self.assertUpdateFetch(
            label2_url, allow_viewers=False, allow_editors=True, form_fields={"name": "Sales", "messages": None}
        )

        # try to update to invalid name
        self.assertUpdateSubmit(
            label1_url,
            {"name": '"Spam"'},
            form_errors={"name": 'Cannot contain the character: "'},
            object_unchanged=label1,
        )

        # update with valid name
        self.assertUpdateSubmit(label1_url, {"name": "Junk"})

        label1.refresh_from_db()
        self.assertEqual("Junk", label1.name)

    def test_delete(self):
        label = self.create_label("Spam")

        delete_url = reverse("msgs.label_delete", args=[label.uuid])

        # fetch delete modal
        response = self.assertDeleteFetch(delete_url, allow_editors=True)
        self.assertContains(response, "You are about to delete")

        # submit to delete it
        response = self.assertDeleteSubmit(delete_url, object_deactivated=label, success_status=200)
        self.assertEqual("/msg/", response["Temba-Success"])

        # reactivate
        label.is_active = True
        label.save()

        # add a dependency and try again
        flow = self.create_flow("Color Flow")
        flow.label_dependencies.add(label)
        self.assertFalse(flow.has_issues)

        response = self.assertDeleteFetch(delete_url, allow_editors=True)
        self.assertContains(response, "is used by the following items but can still be deleted:")
        self.assertContains(response, "Color Flow")

        self.assertDeleteSubmit(delete_url, object_deactivated=label, success_status=200)

        flow.refresh_from_db()
        self.assertTrue(flow.has_issues)
        self.assertNotIn(label, flow.label_dependencies.all())

    def test_list(self):
        self.create_label("Spam")
        self.create_label("Junk")
        self.create_label("Important")
        self.create_label("Other Org", org=self.org2)

        # viewers can't edit flows so don't have access to this JSON endpoint as that's only place it's used
        self.login(self.user)
        response = self.client.get(reverse("msgs.label_list"))
        self.assertLoginRedirect(response)

        # editors can though
        self.login(self.editor)
        response = self.client.get(reverse("msgs.label_list"))
        results = response.json()

        # results should be A-Z and not include folders or labels from other orgs
        self.assertEqual(3, len(results))
        self.assertEqual("Important", results[0]["text"])
        self.assertEqual("Junk", results[1]["text"])
        self.assertEqual("Spam", results[2]["text"])


class SystemLabelTest(TembaTest):
    def test_get_archive_query(self):
        tcs = (
            (
                SystemLabel.TYPE_INBOX,
                "SELECT s.* FROM s3object s WHERE s.direction = 'in' AND s.visibility = 'visible' AND s.status = 'handled' AND s.flow IS NULL AND s.type != 'voice'",
            ),
            (
                SystemLabel.TYPE_FLOWS,
                "SELECT s.* FROM s3object s WHERE s.direction = 'in' AND s.visibility = 'visible' AND s.status = 'handled' AND s.flow IS NOT NULL AND s.type != 'voice'",
            ),
            (
                SystemLabel.TYPE_ARCHIVED,
                "SELECT s.* FROM s3object s WHERE s.direction = 'in' AND s.visibility = 'archived' AND s.status = 'handled' AND s.type != 'voice'",
            ),
            (
                SystemLabel.TYPE_OUTBOX,
                "SELECT s.* FROM s3object s WHERE s.direction = 'out' AND s.visibility = 'visible' AND s.status IN ('initializing', 'queued', 'errored')",
            ),
            (
                SystemLabel.TYPE_SENT,
                "SELECT s.* FROM s3object s WHERE s.direction = 'out' AND s.visibility = 'visible' AND s.status IN ('wired', 'sent', 'delivered')",
            ),
            (
                SystemLabel.TYPE_FAILED,
                "SELECT s.* FROM s3object s WHERE s.direction = 'out' AND s.visibility = 'visible' AND s.status = 'failed'",
            ),
        )

        for label_type, expected_select in tcs:
            select = s3.compile_select(where=SystemLabel.get_archive_query(label_type))
            self.assertEqual(expected_select, select, f"select s3 mismatch for label {label_type}")

    def test_get_counts(self):
        self.assertEqual(
            SystemLabel.get_counts(self.org),
            {
                SystemLabel.TYPE_INBOX: 0,
                SystemLabel.TYPE_FLOWS: 0,
                SystemLabel.TYPE_ARCHIVED: 0,
                SystemLabel.TYPE_OUTBOX: 0,
                SystemLabel.TYPE_SENT: 0,
                SystemLabel.TYPE_FAILED: 0,
                SystemLabel.TYPE_SCHEDULED: 0,
            },
        )

        contact1 = self.create_contact("Bob", phone="0783835001")
        contact2 = self.create_contact("Jim", phone="0783835002")
        msg1 = self.create_incoming_msg(contact1, "Message 1")
        self.create_incoming_msg(contact1, "Message 2")
        msg3 = self.create_incoming_msg(contact1, "Message 3")
        msg4 = self.create_incoming_msg(contact1, "Message 4")
        Broadcast.create(self.org, self.user, {"eng": "Broadcast 2"}, contacts=[contact1, contact2])
        Broadcast.create(
            self.org,
            self.user,
            {"eng": "Broadcast 2"},
            contacts=[contact1, contact2],
            schedule=Schedule.create_schedule(self.org, self.user, timezone.now(), Schedule.REPEAT_DAILY),
        )

        self.assertEqual(
            SystemLabel.get_counts(self.org),
            {
                SystemLabel.TYPE_INBOX: 4,
                SystemLabel.TYPE_FLOWS: 0,
                SystemLabel.TYPE_ARCHIVED: 0,
                SystemLabel.TYPE_OUTBOX: 1,
                SystemLabel.TYPE_SENT: 0,
                SystemLabel.TYPE_FAILED: 0,
                SystemLabel.TYPE_SCHEDULED: 1,
            },
        )

        msg3.archive()

        bcast1 = self.create_broadcast(
            self.user, "Broadcast 1", contacts=[contact1, contact2], msg_status=Msg.STATUS_INITIALIZING
        )
        msg5, msg6 = tuple(Msg.objects.filter(broadcast=bcast1))

        Broadcast.create(
            self.org,
            self.user,
            {"eng": "Broadcast 3"},
            contacts=[contact1],
            schedule=Schedule.create_schedule(self.org, self.user, timezone.now(), Schedule.REPEAT_DAILY),
        )

        self.assertEqual(
            SystemLabel.get_counts(self.org),
            {
                SystemLabel.TYPE_INBOX: 3,
                SystemLabel.TYPE_FLOWS: 0,
                SystemLabel.TYPE_ARCHIVED: 1,
                SystemLabel.TYPE_OUTBOX: 3,
                SystemLabel.TYPE_SENT: 0,
                SystemLabel.TYPE_FAILED: 0,
                SystemLabel.TYPE_SCHEDULED: 2,
            },
        )

        msg1.archive()
        msg3.delete()  # deleting an archived msg
        msg4.delete()  # deleting a visible msg
        msg5.status = "F"
        msg5.save(update_fields=("status",))
        msg6.status = "S"
        msg6.save(update_fields=("status",))

        self.assertEqual(
            SystemLabel.get_counts(self.org),
            {
                SystemLabel.TYPE_INBOX: 1,
                SystemLabel.TYPE_FLOWS: 0,
                SystemLabel.TYPE_ARCHIVED: 1,
                SystemLabel.TYPE_OUTBOX: 1,
                SystemLabel.TYPE_SENT: 1,
                SystemLabel.TYPE_FAILED: 1,
                SystemLabel.TYPE_SCHEDULED: 2,
            },
        )

        msg1.restore()
        msg5.status = "F"  # already failed
        msg5.save(update_fields=("status",))
        msg6.status = "D"
        msg6.save(update_fields=("status",))

        self.assertEqual(
            SystemLabel.get_counts(self.org),
            {
                SystemLabel.TYPE_INBOX: 2,
                SystemLabel.TYPE_FLOWS: 0,
                SystemLabel.TYPE_ARCHIVED: 0,
                SystemLabel.TYPE_OUTBOX: 1,
                SystemLabel.TYPE_SENT: 1,
                SystemLabel.TYPE_FAILED: 1,
                SystemLabel.TYPE_SCHEDULED: 2,
            },
        )

        self.assertEqual(SystemLabelCount.objects.all().count(), 21)

        # squash our counts
        squash_msg_counts()

        self.assertEqual(
            SystemLabel.get_counts(self.org),
            {
                SystemLabel.TYPE_INBOX: 2,
                SystemLabel.TYPE_FLOWS: 0,
                SystemLabel.TYPE_ARCHIVED: 0,
                SystemLabel.TYPE_OUTBOX: 1,
                SystemLabel.TYPE_SENT: 1,
                SystemLabel.TYPE_FAILED: 1,
                SystemLabel.TYPE_SCHEDULED: 2,
            },
        )

        # we should only have one system label per type
        self.assertEqual(SystemLabelCount.objects.all().count(), 6)


class TagsTest(TembaTest):
    def setUp(self):
        super().setUp()

        self.joe = self.create_contact("Joe Blow", phone="+250788382382")

    def render_template(self, string, context=None):
        from django.template import Context, Template

        context = context or {}
        context = Context(context)
        return Template(string).render(context)

    def assertHasClass(self, text, clazz):
        self.assertTrue(text.find(clazz) >= 0)

    def test_as_icon(self):
        msg = self.create_outgoing_msg(self.joe, "How is it going?", status=Msg.STATUS_QUEUED)
        now = timezone.now()
        two_hours_ago = now - timedelta(hours=2)

        self.assertHasClass(as_icon(msg), "icon-bubble-dots-2 green")
        msg.created_on = two_hours_ago
        self.assertHasClass(as_icon(msg), "icon-bubble-dots-2 green")
        msg.status = "S"
        self.assertHasClass(as_icon(msg), "icon-bubble-right green")
        msg.status = "D"
        self.assertHasClass(as_icon(msg), "icon-bubble-check green")
        msg.status = "E"
        self.assertHasClass(as_icon(msg), "icon-bubble-notification red")
        msg.direction = "I"
        self.assertHasClass(as_icon(msg), "icon-bubble-user primary")
        msg.msg_type = "V"
        self.assertHasClass(as_icon(msg), "icon-phone")

        # default cause is pending sent
        self.assertHasClass(as_icon(None), "icon-bubble-dots-2 green")

        in_call = self.create_channel_event(
            self.channel, str(self.joe.get_urn(URN.TEL_SCHEME)), ChannelEvent.TYPE_CALL_IN
        )
        self.assertHasClass(as_icon(in_call), "icon-call-incoming green")

        in_miss = self.create_channel_event(
            self.channel, str(self.joe.get_urn(URN.TEL_SCHEME)), ChannelEvent.TYPE_CALL_IN_MISSED
        )
        self.assertHasClass(as_icon(in_miss), "icon-call-incoming red")

        out_call = self.create_channel_event(
            self.channel, str(self.joe.get_urn(URN.TEL_SCHEME)), ChannelEvent.TYPE_CALL_OUT
        )
        self.assertHasClass(as_icon(out_call), "icon-call-outgoing green")

        out_miss = self.create_channel_event(
            self.channel, str(self.joe.get_urn(URN.TEL_SCHEME)), ChannelEvent.TYPE_CALL_OUT_MISSED
        )
        self.assertHasClass(as_icon(out_miss), "icon-call-outgoing red")

    def test_render(self):
        template_src = "{% load sms %}{% render as foo %}123<a>{{ bar }}{% endrender %}-{{ foo }}-"
        self.assertEqual(self.render_template(template_src, {"bar": "abc"}), "-123<a>abc-")

        # exception if tag not used correctly
        self.assertRaises(ValueError, self.render_template, "{% load sms %}{% render with bob %}{% endrender %}")
        self.assertRaises(ValueError, self.render_template, "{% load sms %}{% render as %}{% endrender %}")


class MediaCRUDLTest(CRUDLTestMixin, TembaTest):
    @mock_uuids
    def test_upload(self):
        upload_url = reverse("msgs.media_upload")

        def assert_upload(user, filename, expected_json):
            self.login(user)

            with open(filename, "rb") as data:
                response = self.client.post(upload_url, {"file": data}, HTTP_X_FORWARDED_HTTPS="https")

            self.assertEqual(response.status_code, 200)
            self.assertEqual(expected_json, response.json())

        assert_upload(
            self.admin,
            f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg",
            {
                "uuid": "b97f69f7-5edf-45c7-9fda-d37066eae91d",
                "content_type": "image/jpeg",
                "type": "image/jpeg",
                "url": f"/media/test_orgs/{self.org.id}/media/b97f/b97f69f7-5edf-45c7-9fda-d37066eae91d/steve%20marten.jpg",
                "name": "steve marten.jpg",
                "size": 7461,
            },
        )
        assert_upload(
            self.editor,
            f"{settings.MEDIA_ROOT}/test_media/snow.mp4",
            {
                "uuid": "14f6ea01-456b-4417-b0b8-35e942f549f1",
                "content_type": "video/mp4",
                "type": "video/mp4",
                "url": f"/media/test_orgs/{self.org.id}/media/14f6/14f6ea01-456b-4417-b0b8-35e942f549f1/snow.mp4",
                "name": "snow.mp4",
                "size": 684558,
            },
        )
        assert_upload(
            self.editor,
            f"{settings.MEDIA_ROOT}/test_media/bubbles.m4a",
            {
                "uuid": "9295ebab-5c2d-4eb1-86f9-7c15ed2f3219",
                "content_type": "audio/mp4",
                "type": "audio/mp4",
                "url": f"/media/test_orgs/{self.org.id}/media/9295/9295ebab-5c2d-4eb1-86f9-7c15ed2f3219/bubbles.m4a",
                "name": "bubbles.m4a",
                "size": 46468,
            },
        )

        # error message if you upload something unsupported
        with open(f"{settings.MEDIA_ROOT}/test_imports/simple.xls", "rb") as data:
            response = self.client.post(upload_url, {"file": data}, HTTP_X_FORWARDED_HTTPS="https")
            self.assertEqual({"error": "Unsupported file type"}, response.json())

        # error message if upload is too big
        with patch("temba.msgs.models.Media.MAX_UPLOAD_SIZE", 1024):
            with open(f"{settings.MEDIA_ROOT}/test_media/snow.mp4", "rb") as data:
                response = self.client.post(upload_url, {"file": data}, HTTP_X_FORWARDED_HTTPS="https")
                self.assertEqual({"error": "Limit for file uploads is 0.0009765625 MB"}, response.json())

        self.clear_storage()

    def test_list(self):
        upload_url = reverse("msgs.media_upload")
        list_url = reverse("msgs.media_list")

        def upload(user, path):
            self.login(user)

            with open(path, "rb") as data:
                self.client.post(upload_url, {"file": data}, HTTP_X_FORWARDED_HTTPS="https")
                return self.org.media.filter(original=None).order_by("id").last()

        media1 = upload(self.admin, f"{settings.MEDIA_ROOT}/test_media/steve marten.jpg")
        media2 = upload(self.admin, f"{settings.MEDIA_ROOT}/test_media/bubbles.m4a")
        upload(self.admin2, f"{settings.MEDIA_ROOT}/test_media/bubbles.m4a")  # other org

        self.login(self.customer_support, choose_org=self.org)
        response = self.client.get(list_url)
        self.assertEqual([media2, media1], list(response.context["object_list"]))

        self.clear_storage()
